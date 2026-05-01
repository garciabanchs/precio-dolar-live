from fastapi.responses import FileResponse, JSONResponse, HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi import FastAPI, UploadFile, File, Form, Body, HTTPException, Request, Response
from pathlib import Path
from fx_extractors import build_flat_fx_values
import json
import csv
import zipfile
import os
import hashlib
import uuid
from dotenv import load_dotenv
try:
    import stripe
except ImportError:
    stripe = None
from datetime import datetime, timezone, timedelta
from html import escape

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
    PageBreak,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.units import mm

from backend.services.excel_reader import read_excel_payload
from backend.services.pricing_engine import apply_pricing_engine
from backend.services.fx_service import (
    get_real_fx_snapshot,
    get_reference_pair,
    get_fx_summary,
    get_partial_real_fx_sources
)
from backend.services.fx_orchestrator import (
    build_fx_check_response,
    validate_manual_completion,
    build_fx_complete_response,
    to_pricing_snapshot
)

from backend.services.fx_sheet_cache_service import (
    get_pricing_fx_context,
    get_fx_history_cached
)

import requests

CRONSECRET = os.getenv("CRONSECRET")
if not CRONSECRET:
    raise RuntimeError("CRONSECRET no configurado en entorno")

SHEET_API = "https://script.google.com/macros/s/AKfycbyYmWspNw9yX06pqUEfNc3nnj2UdX8jQrdI5PZ1DBzAKzGDXpf5sKllp6S-YMQU-yIzqQ/exec"

app = FastAPI()

ALLOWED_ORIGINS = [
    origin.strip()
    for origin in os.getenv(
        "ALLOWED_ORIGINS",
        "http://127.0.0.1:8000,http://localhost:8000"
    ).split(",")
    if origin.strip()
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST"],
    allow_headers=["Content-Type", "Authorization"],
)

BASE_DIR = Path(__file__).resolve().parent
ASSETS_DIR = BASE_DIR / "frontend" / "assets"
DATA_DIR = BASE_DIR / "data"

app.mount("/frontend", StaticFiles(directory=str(BASE_DIR / "frontend")), name="frontend")

FX_HISTORY_PATH = DATA_DIR / "fx_history.json"
REPORTS_DIR = DATA_DIR / "reports"
PDF_DIR = DATA_DIR / "pdfs"
EXCEL_DIR = DATA_DIR / "excels"
HTML_DIR = DATA_DIR / "html"
ZIP_DIR = DATA_DIR / "zips"
ECOMMERCE_DIR = DATA_DIR / "ecommerce"
ACCESS_CONTROL_PATH = DATA_DIR / "access_control.json"
ACCESS_COOKIE_NAME = "precio_dolar_client_id"
FREE_ACCESS_DAYS = 15
MIN_FREE_UNIQUE_USES = 3
CLIENT_COOKIE_MAX_AGE_SECONDS = 60 * 60 * 24 * 365
load_dotenv()
IS_PROD = os.getenv("ENV") == "production"
ADMIN_TOKEN = os.getenv("ADMIN_TOKEN")
if not ADMIN_TOKEN:
    raise RuntimeError("ADMIN_TOKEN no configurado en entorno")

ACCESS_EMAIL_COOKIE_NAME = "precio_dolar_user_email"
ADMIN_SESSION_COOKIE_NAME = "precio_dolar_admin_session"
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD")
if not ADMIN_USERNAME or not ADMIN_PASSWORD:
    raise RuntimeError("ADMIN_USERNAME y ADMIN_PASSWORD deben configurarse")
ADMIN_SESSION_VALUE = os.getenv("ADMIN_SESSION_VALUE") or hashlib.sha256(f"{ADMIN_USERNAME}:{ADMIN_PASSWORD}:{ADMIN_TOKEN}".encode("utf-8")).hexdigest()
VALID_ACCESS_PLANS = {"basic", "premium"}

STRIPE_SECRET_KEY = os.getenv("STRIPE_SECRET_KEY")
STRIPE_WEBHOOK_SECRET = os.getenv("STRIPE_WEBHOOK_SECRET")
STRIPE_PRICE_BASIC = os.getenv("STRIPE_PRICE_BASIC")
STRIPE_PRICE_PREMIUM = os.getenv("STRIPE_PRICE_PREMIUM")
PUBLIC_BASE_URL = os.getenv("PUBLIC_BASE_URL")
if PUBLIC_BASE_URL:
    PUBLIC_BASE_URL = PUBLIC_BASE_URL.rstrip("/")

if stripe is not None:
    stripe.api_key = STRIPE_SECRET_KEY

VALID_REPORT_TYPES = {"ejecutivo", "operativo", "tienda"}
VALID_FX_KEYS = {"compuesto", "monitor", "binance", "usdt", "dolartoday"}

def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def parse_iso_datetime(value: str):
    try:
        dt = datetime.fromisoformat(value)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        return None
    
def format_admin_datetime(value):
    if not value:
        return "-"

    try:
        dt = parse_iso_datetime(str(value))
        if dt is None:
            return str(value)
        return dt.astimezone().strftime("%d/%m/%Y %I:%M %p")
    except Exception:
        return str(value)


def hash_uploaded_file(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def normalize_price_for_hash(value) -> str:
    try:
        if value is None or value == "":
            return "0.000000"
        return f"{float(value):.6f}"
    except Exception:
        return str(value).strip()


def build_old_prices_hash_from_payload(payload: dict) -> str:
    """
    Construye un hash estable basado SOLO en los precios viejos del archivo cargado.
    No depende del nombre del archivo, formato, estilos, fórmulas, filtros ni metadatos del Excel.
    """
    price_values = []

    for mercado in payload.get("mercados", []) or []:
        df = mercado.get("productos")
        if df is None:
            continue

        candidate_columns = [
            "precio_propio_usd",
            "precio_viejo_usd",
            "precio_actual_usd",
            "precio_anterior_usd",
            "precio_propio",
            "precio_viejo",
            "precio_actual",
            "precio_anterior",
        ]

        selected_column = None
        for column in candidate_columns:
            if column in df.columns:
                selected_column = column
                break

        if selected_column is None:
            for column in df.columns:
                column_text = str(column).lower()
                if "precio" in column_text and ("propio" in column_text or "viejo" in column_text or "actual" in column_text or "anterior" in column_text):
                    selected_column = column
                    break

        if selected_column is None:
            continue

        for value in df[selected_column].tolist():
            price_values.append(normalize_price_for_hash(value))

    serialized = json.dumps(price_values, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()


def ensure_data_dir():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    PDF_DIR.mkdir(parents=True, exist_ok=True)
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    HTML_DIR.mkdir(parents=True, exist_ok=True)
    ZIP_DIR.mkdir(parents=True, exist_ok=True)
    ECOMMERCE_DIR.mkdir(parents=True, exist_ok=True)


def get_client_report_dir(client_id: str) -> Path:
    ensure_data_dir()
    safe_client_id = sanitize_filename(str(client_id))
    client_dir = REPORTS_DIR / safe_client_id
    client_dir.mkdir(parents=True, exist_ok=True)
    return client_dir


def get_client_last_report_path(client_id: str) -> Path:
    return get_client_report_dir(client_id) / "last_report.json"

def get_email_report_dir(email: str) -> Path:
    ensure_data_dir()
    safe_email = sanitize_filename(str(email))
    email_dir = REPORTS_DIR / "email" / safe_email
    email_dir.mkdir(parents=True, exist_ok=True)
    return email_dir


def get_email_last_report_path(email: str) -> Path:
    return get_email_report_dir(email) / "last_report.json"


def merge_normalized_access_users(users: dict) -> dict:
    if not isinstance(users, dict):
        return {}

    merged = {}

    for raw_email, raw_user in users.items():
        normalized_email = normalize_email(raw_email)
        if not normalized_email:
            continue

        candidate = normalize_access_user(normalized_email, raw_user)
        current = merged.get(normalized_email)

        if current is None:
            merged[normalized_email] = candidate
            continue

        candidate_active = is_user_access_active(candidate)
        current_active = is_user_access_active(current)

        candidate_until = parse_iso_datetime(candidate.get("access_until")) if candidate.get("access_until") else None
        current_until = parse_iso_datetime(current.get("access_until")) if current.get("access_until") else None

        prefer_candidate = False

        if candidate_active and not current_active:
            prefer_candidate = True
        elif candidate_active == current_active:
            if candidate_until and not current_until:
                prefer_candidate = True
            elif candidate_until and current_until and candidate_until > current_until:
                prefer_candidate = True

        if prefer_candidate:
            merged[normalized_email] = candidate

    return merged


def load_access_control() -> dict:
    ensure_data_dir()

    if not ACCESS_CONTROL_PATH.exists():
        return {"clients": {}, "users": {}, "client_email_map": {}, "stripe_events": []}

    try:
        with open(ACCESS_CONTROL_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return {"clients": {}, "users": {}, "client_email_map": {}, "stripe_events": []}

    if not isinstance(data, dict):
        return {"clients": {}, "users": {}, "client_email_map": {}, "stripe_events": []}

    if not isinstance(data.get("clients"), dict):
        data["clients"] = {}
    if not isinstance(data.get("users"), dict):
        data["users"] = {}
    if not isinstance(data.get("client_email_map"), dict):
        data["client_email_map"] = {}
    if not isinstance(data.get("stripe_events"), list):
        data["stripe_events"] = []

    original_users = data.get("users", {})
    normalized_users = merge_normalized_access_users(original_users)

    if normalized_users != original_users:
        data["users"] = normalized_users
        save_access_control(data)
    else:
        data["users"] = normalized_users

    return data



def save_access_control(data: dict) -> None:
    ensure_data_dir()
    with open(ACCESS_CONTROL_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_or_create_client_id(request: Request, response: Response) -> str:
    client_id = request.cookies.get(ACCESS_COOKIE_NAME)

    if not client_id:
        client_id = str(uuid.uuid4())

    response.set_cookie(
        key=ACCESS_COOKIE_NAME,
        value=client_id,
        max_age=CLIENT_COOKIE_MAX_AGE_SECONDS,
        httponly=True,
        samesite="lax",
        secure=IS_PROD
    )
    return client_id


def get_client_id_from_request(request: Request) -> str | None:
    return request.cookies.get(ACCESS_COOKIE_NAME)


def get_or_create_access_client(data: dict, client_id: str) -> dict:
    clients = data.setdefault("clients", {})
    client = clients.get(client_id)
    now = utc_now_iso()

    if not isinstance(client, dict):
        client = {}
        clients[client_id] = client

    legacy_first_use_at = client.get("first_use_at")
    legacy_hashes = client.get("unique_file_hashes")
    legacy_usage_count = client.get("usage_count")

    client.setdefault("client_id", client_id)
    client.setdefault("created_at", legacy_first_use_at or now)
    client.setdefault("uploads_count", legacy_usage_count or 0)
    client.setdefault("distinct_price_hashes", legacy_hashes if isinstance(legacy_hashes, list) else [])
    client.setdefault("last_upload_at", None)
    client.setdefault("is_paid_basic", False)
    client.setdefault("is_paid_ecommerce", False)

    if not isinstance(client.get("distinct_price_hashes"), list):
        client["distinct_price_hashes"] = []

    try:
        client["uploads_count"] = int(client.get("uploads_count", 0) or 0)
    except Exception:
        client["uploads_count"] = len(client["distinct_price_hashes"])

    client["distinct_files_count"] = len(client["distinct_price_hashes"])
    client["first_use_at"] = client.get("created_at")
    client["usage_count"] = len(client["distinct_price_hashes"])
    client["unique_file_hashes"] = client["distinct_price_hashes"]

    return client


def get_access_client(client_id: str) -> dict:
    data = load_access_control()
    client = get_or_create_access_client(data, client_id)
    save_access_control(data)
    return client


def build_access_status(client: dict) -> dict:
    created_at = client.get("created_at")
    created_dt = parse_iso_datetime(created_at) if created_at else None
    now = datetime.now(timezone.utc)

    days_since_created = 0
    trial_days_remaining = FREE_ACCESS_DAYS

    if created_dt is not None:
        days_since_created = max((now - created_dt).days, 0)
        trial_days_remaining = max(FREE_ACCESS_DAYS - days_since_created, 0)

    distinct_count = len(client.get("distinct_price_hashes", []) or [])
    uploads_count = int(client.get("uploads_count", 0) or 0)
    unique_uses_remaining = max(MIN_FREE_UNIQUE_USES - distinct_count, 0)
    blocked = days_since_created >= FREE_ACCESS_DAYS and distinct_count >= MIN_FREE_UNIQUE_USES

    return {
        "client_id": client.get("client_id"),
        "created_at": created_at,
        "first_use_at": created_at,
        "uploads_count": uploads_count,
        "distinct_files_count": distinct_count,
        "usage_count": distinct_count,
        "free_access_days": FREE_ACCESS_DAYS,
        "min_free_unique_uses": MIN_FREE_UNIQUE_USES,
        "days_since_created": days_since_created,
        "days_since_first_use": days_since_created,
        "trial_days_remaining": trial_days_remaining,
        "unique_uses_remaining": unique_uses_remaining,
        "last_upload_at": client.get("last_upload_at"),
        "is_paid_basic": bool(client.get("is_paid_basic", False)),
        "is_paid_ecommerce": bool(client.get("is_paid_ecommerce", False)),
        "blocked": blocked
    }


def is_blocked(client: dict) -> bool:
    return build_access_status(client).get("blocked", False)


def trial_expired_response() -> JSONResponse:
    return JSONResponse(
        status_code=402,
        content={
            "blocked": True,
            "reason": "trial_expired",
            "message": "Tu período de prueba ha finalizado. Elige un plan para continuar.",
            "payment_options": {
                "basic": "Plan operativo",
                "ecommerce": "Plan e-commerce"
            }
        }
    )


def normalize_email(value: str | None) -> str | None:
    if value is None:
        return None
    email = str(value).strip().lower()
    if not email or "@" not in email:
        return None
    return email


def calculate_access_until(duration_days=None, duration: str | None = None) -> str:
    now = datetime.now(timezone.utc)
    if duration_days is not None:
        try:
            days = int(duration_days)
        except Exception:
            days = 30
        return (now + timedelta(days=max(days, 1))).isoformat()

    normalized = str(duration or "30d").strip().lower()
    if normalized in {"1y", "year", "yearly", "annual", "anual", "+1 año", "+1 ano"}:
        return (now + timedelta(days=365)).isoformat()

    try:
        if normalized.endswith("d"):
            return (now + timedelta(days=max(int(normalized[:-1]), 1))).isoformat()
        if normalized.endswith("y"):
            return (now + timedelta(days=max(int(normalized[:-1]), 1) * 365)).isoformat()
    except Exception:
        pass

    return (now + timedelta(days=30)).isoformat()


def get_email_from_request(request: Request) -> str | None:
    email = normalize_email(request.query_params.get("email"))
    if email:
        return email

    email = normalize_email(request.headers.get("X-User-Email"))
    if email:
        return email

    email = normalize_email(request.cookies.get(ACCESS_EMAIL_COOKIE_NAME))
    if email:
        return email

    client_id = get_client_id_from_request(request)
    if client_id:
        data = load_access_control()
        return normalize_email(data.get("client_email_map", {}).get(client_id))

    return None


def bind_email_to_client(request: Request, response: Response, email: str) -> None:
    client_id = get_or_create_client_id(request, response)
    data = load_access_control()
    data.setdefault("client_email_map", {})[client_id] = email
    save_access_control(data)
    response.set_cookie(
        key=ACCESS_EMAIL_COOKIE_NAME,
        value=email,
        max_age=CLIENT_COOKIE_MAX_AGE_SECONDS,
        httponly=True,
        samesite="lax",
        secure=IS_PROD
    )


def normalize_access_user(email: str, user: dict | None = None, full_name: str | None = None) -> dict:
    now = utc_now_iso()
    if not isinstance(user, dict):
        user = {}

    user.setdefault("email", email)
    user.setdefault("plan", "basic")
    user.setdefault("access_until", None)
    user.setdefault("active", False)
    user.setdefault("created_at", now)
    user.setdefault("updated_at", now)
    user.setdefault("source", "manual")
    user.setdefault("stripe_customer_id", None)
    user.setdefault("stripe_subscription_id", None)
    user.setdefault("full_name", None)

    if user.get("plan") not in VALID_ACCESS_PLANS:
        user["plan"] = "basic"

    user["email"] = email
    user["active"] = bool(user.get("active", False))

    if full_name:
        user["full_name"] = full_name
    elif not user.get("full_name"):
        user["full_name"] = None
    return user


def get_access_user(email: str | None) -> dict | None:
    normalized_email = normalize_email(email)
    if not normalized_email:
        return None

    data = load_access_control()
    user = data.get("users", {}).get(normalized_email)
    if not isinstance(user, dict):
        return None

    user = normalize_access_user(normalized_email, user)
    data.setdefault("users", {})[normalized_email] = user
    save_access_control(data)
    return user


def is_user_access_active(user: dict | None) -> bool:
    if not isinstance(user, dict) or not bool(user.get("active", False)):
        return False

    access_until = parse_iso_datetime(user.get("access_until")) if user.get("access_until") else None
    return bool(access_until and access_until >= datetime.now(timezone.utc))


def build_user_access_status(user: dict | None) -> dict:
    if not isinstance(user, dict):
        return {
            "has_access": False,
            "reason": "user_not_found",
            "plan": None,
            "email": None,
            "access_until": None,
            "active": False
        }

    active = is_user_access_active(user)
    return {
        "has_access": active,
        "reason": None if active else "access_inactive_or_expired",
        "plan": user.get("plan"),
        "email": user.get("email"),
        "access_until": user.get("access_until"),
        "active": bool(user.get("active", False))
    }


def access_denied_response(reason: str, required_plan: str = "basic") -> JSONResponse:
    return JSONResponse(
        status_code=402,
        content={
            "blocked": True,
            "reason": reason,
            "required_plan": required_plan,
            "message": "No tienes acceso activo para esta descarga. Inicia sesión con un email autorizado o solicita activación."
        }
    )


def require_download_access(request: Request, requires_premium: bool = False):
    email = get_email_from_request(request)
    user = get_access_user(email)

    if user is not None:
        if not is_user_access_active(user):
            return access_denied_response(
                "access_inactive_or_expired",
                "premium" if requires_premium else "basic"
            )

        plan = str(user.get("plan") or "").strip().lower()

        if requires_premium and plan != "premium":
            return access_denied_response("premium_required", "premium")

        return None

    client_id = get_client_id_from_request(request)
    if client_id:
        client = get_access_client(client_id)

        if not is_blocked(client):
            return None

        if requires_premium:
            return access_denied_response("premium_required", "premium")

        if bool(client.get("is_paid_basic", False)) or bool(client.get("is_paid_ecommerce", False)):
            return None

    return access_denied_response("access_required", "premium" if requires_premium else "basic")

def upsert_access_user(email: str, plan: str = "basic", access_until: str | None = None, active: bool = True, source: str = "manual", full_name: str | None = None) -> dict:
    normalized_email = normalize_email(email)
    if not normalized_email:
        raise HTTPException(status_code=400, detail="Email inválido u obligatorio.")
    if plan not in VALID_ACCESS_PLANS:
        raise HTTPException(status_code=400, detail="Plan inválido. Usa 'basic' o 'premium'.")

    data = load_access_control()
    users = data.setdefault("users", {})
    existing_user = users.get(normalized_email)
    user = normalize_access_user(normalized_email, existing_user, full_name=full_name)

    existing_source = user.get("source") if isinstance(existing_user, dict) else None

    user["plan"] = plan
    user["access_until"] = access_until or calculate_access_until(duration="30d")
    user["active"] = bool(active)
    user["source"] = "stripe" if source == "stripe" else (existing_source or source)
    user["updated_at"] = utc_now_iso()

    users[normalized_email] = user
    save_access_control(data)
    return user



def set_user_active(email: str, active: bool) -> dict:
    normalized_email = normalize_email(email)
    if not normalized_email:
        raise HTTPException(status_code=400, detail="Email inválido u obligatorio.")

    data = load_access_control()
    users = data.setdefault("users", {})
    if normalized_email not in users:
        raise HTTPException(status_code=404, detail="Usuario no encontrado.")

    user = normalize_access_user(normalized_email, users[normalized_email])
    user["active"] = bool(active)
    user["updated_at"] = utc_now_iso()
    users[normalized_email] = user
    save_access_control(data)
    return user


def require_admin_session(request: Request) -> None:
    if request.cookies.get(ADMIN_SESSION_COOKIE_NAME) == ADMIN_SESSION_VALUE:
        return
    raise HTTPException(status_code=401, detail="Sesión administrativa inválida.")


def admin_is_logged_in(request: Request) -> bool:
    return request.cookies.get(ADMIN_SESSION_COOKIE_NAME) == ADMIN_SESSION_VALUE


def render_admin_page(request: Request, message: str = "") -> HTMLResponse:
    if not admin_is_logged_in(request):
        return HTMLResponse(f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
                            <link rel="icon" href="data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'%3E%3Ctext y='.9em' font-size='90'%3E🧠%3C/text%3E%3C/svg%3E">
<title>Panel admin · PrecioDólar LIVE</title>
<style>
body{{font-family:Arial,Helvetica,sans-serif;background:#f8fafc;color:#0f172a;margin:0;padding:32px}}
.card{{max-width:420px;margin:60px auto;background:#fff;border:1px solid #cbd5e1;border-radius:14px;padding:24px}}
label{{display:block;margin-top:12px;font-weight:bold}}
input{{width:100%;padding:10px;margin-top:6px;border:1px solid #cbd5e1;border-radius:8px;box-sizing:border-box}}
button{{margin-top:18px;background:#064e3b;color:#fff;border:0;border-radius:8px;padding:10px 14px;font-weight:bold;cursor:pointer}}
.msg{{color:#b91c1c}}
</style></head><body>
<div class="card"><h1>Panel admin · PrecioDólar LIVE</h1><p class="msg">{escape(message)}</p>
<form method="post" action="/admin/login">
<label>Usuario</label><input name="username" autocomplete="username" required>
<label>Contraseña</label><input name="password" type="password" autocomplete="current-password" required>
<button type="submit">Entrar</button></form></div></body></html>""")

    data = load_access_control()
    raw_users = data.get("users", {})

    if isinstance(raw_users, dict):
        users = [normalize_access_user(email, user) for email, user in raw_users.items()]
    elif isinstance(raw_users, list):
        users = [normalize_access_user(user.get("email"), user) for user in raw_users if isinstance(user, dict)]
    else:
        users = []

    users = [user for user in users if user.get("email")]
    users.sort(key=lambda item: item.get("email", ""))

    rows = ""
    for user in users:
        status = "Inactivo"
        if user.get("active"):
            if is_user_access_active(user):
                status = "Activo"
            else:
                status = "Expirado"

        rows += f"""<tr>
<td>{escape(str(user.get('full_name') or '-'))}</td>
<td>{escape(str(user.get('email') or ''))}</td>
<td>{escape(format_admin_datetime(user.get('created_at')))}</td>
<td>{escape(str(user.get('plan') or ''))}</td>
<td>{escape(format_admin_datetime(user.get('access_until')))}</td>
<td>{escape(status)}</td>
<td>{escape(str(user.get('source') or 'manual'))}</td>
<td><form method="post" action="/admin/users/upsert" class="inline">
<input type="hidden" name="email" value="{escape(user.get('email', ''))}">
<select name="plan"><option value="basic" {'selected' if user.get('plan') == 'basic' else ''}>basic</option><option value="premium" {'selected' if user.get('plan') == 'premium' else ''}>premium</option></select>
<input name="duration_days" type="number" min="1" value="30" style="width:80px">
<select name="active"><option value="true" {'selected' if user.get('active') else ''}>activo</option><option value="false" {'selected' if not user.get('active') else ''}>inactivo</option></select>
<button>Guardar</button></form></td></tr>"""

    if not rows:
        rows = '<tr><td colspan="8">No hay usuarios todavía.</td></tr>'

    return HTMLResponse(f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<link rel="icon" href="data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'%3E%3Ctext y='.9em' font-size='90'%3E🧠%3C/text%3E%3C/svg%3E">
<title>Panel admin · PrecioDólar LIVE</title>
<style>
body{{font-family:Arial,Helvetica,sans-serif;background:#f8fafc;color:#0f172a;margin:0;padding:24px}}
.container{{max-width:1180px;margin:0 auto}}
.card{{background:#fff;border:1px solid #cbd5e1;border-radius:14px;padding:18px;margin-bottom:18px}}
h1,h2{{color:#064e3b}} label{{display:block;margin-top:10px;font-weight:bold}}
input,select{{padding:8px;margin-top:5px;border:1px solid #cbd5e1;border-radius:8px}}
button{{background:#064e3b;color:#fff;border:0;border-radius:8px;padding:9px 12px;font-weight:bold;cursor:pointer}}
table{{width:100%;border-collapse:collapse;background:#fff}} th{{background:#064e3b;color:#fff;text-align:left;padding:10px}} td{{border:1px solid #e2e8f0;padding:9px;vertical-align:top}}
.inline{{display:flex;gap:6px;align-items:center;flex-wrap:wrap}} .top{{display:flex;justify-content:space-between;align-items:center}} .msg{{color:#064e3b;font-weight:bold}}
</style></head><body><div class="container">
<div class="top"><h1>Panel admin · PrecioDólar LIVE</h1><form method="post" action="/admin/logout"><button>Salir</button></form></div>
<p class="msg">{escape(message)}</p>
<div class="card"><h2>Crear / activar usuario</h2>
<form method="post" action="/admin/users/upsert" class="inline">
<label>Nombre<br><input name="full_name" type="text"></label>
<label>Email<br><input name="email" type="email" required></label>
<label>Plan<br><select name="plan"><option value="basic">basic</option><option value="premium">premium</option></select></label><label>Duración (días)<br><input name="duration_days" type="number" min="1" value="30"></label>
<label>Estado<br><select name="active"><option value="true">activo</option><option value="false">inactivo</option></select></label>
<button>Guardar</button></form></div>
<div class="card"><h2>Usuarios</h2><table><thead><tr><th>Nombre</th><th>Email</th><th>Creado</th><th>Plan</th><th>Acceso hasta</th><th>Estado</th><th>Fuente</th><th>Editar</th></tr></thead><tbody>{rows}</tbody></table></div>
</div></body></html>""")



def render_access_login_page(message: str = "") -> HTMLResponse:
    return HTMLResponse(f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Acceso · PrecioDólar LIVE</title>
<style>body{{font-family:Arial,Helvetica,sans-serif;background:#f8fafc;color:#0f172a;padding:32px}}.card{{max-width:460px;margin:60px auto;background:#fff;border:1px solid #cbd5e1;border-radius:14px;padding:24px}}input{{width:100%;padding:10px;border:1px solid #cbd5e1;border-radius:8px;box-sizing:border-box}}button{{margin-top:14px;background:#064e3b;color:white;border:0;border-radius:8px;padding:10px 14px;font-weight:bold}}</style>
</head><body><div class="card"><h1>Acceso</h1><p>{escape(message or 'Introduce tu email autorizado para habilitar las descargas en este navegador.')}</p>
<form method="post" action="/access/login"><input name="email" type="email" required><button>Entrar</button></form></div></body></html>""")


def require_basic_access(request: Request):
    return require_download_access(request, requires_premium=False)


def require_ecommerce_access(request: Request):
    return require_download_access(request, requires_premium=True)


def register_unique_file_usage(client_id: str, filename: str, price_hash: str) -> dict:
    data = load_access_control()
    client = get_or_create_access_client(data, client_id)
    now = utc_now_iso()
    counted_as_new_use = False

    client["uploads_count"] = int(client.get("uploads_count", 0) or 0) + 1
    client["last_upload_at"] = now

    if price_hash not in client["distinct_price_hashes"]:
        client["distinct_price_hashes"].append(price_hash)
        counted_as_new_use = True

    client["distinct_files_count"] = len(client["distinct_price_hashes"])
    client["usage_count"] = len(client["distinct_price_hashes"])
    client["unique_file_hashes"] = client["distinct_price_hashes"]
    client["first_use_at"] = client.get("created_at")

    save_access_control(data)

    status = build_access_status(client)
    status["file_hash"] = price_hash
    status["price_hash"] = price_hash
    status["hash_scope"] = "old_prices_only"
    status["counted_as_new_use"] = counted_as_new_use

    return status


def mark_client_paid(client_id: str, plan: str, paid: bool = True) -> dict:
    data = load_access_control()
    client = get_or_create_access_client(data, client_id)

    if plan == "basic":
        client["is_paid_basic"] = bool(paid)
    elif plan == "ecommerce":
        if paid:
            client["is_paid_basic"] = True
        client["is_paid_ecommerce"] = bool(paid)
    else:
        raise HTTPException(status_code=400, detail="Plan inválido. Usa 'basic' o 'ecommerce'.")

    save_access_control(data)
    return build_access_status(client)


def require_admin_token(request: Request) -> None:
    token = request.headers.get("X-Admin-Token")
    if token != ADMIN_TOKEN:
        raise HTTPException(status_code=401, detail="Token administrativo inválido.")
    

def save_last_report(report_data: dict, client_id: str | None = None, email: str | None = None) -> None:
    ensure_data_dir()

    if client_id:
        client_report_path = get_client_last_report_path(client_id)
        with open(client_report_path, "w", encoding="utf-8") as f:
            json.dump(report_data, f, ensure_ascii=False, indent=2)

    normalized_email = normalize_email(email)
    if normalized_email:
        email_report_path = get_email_last_report_path(normalized_email)
        with open(email_report_path, "w", encoding="utf-8") as f:
            json.dump(report_data, f, ensure_ascii=False, indent=2)


def load_last_report(client_id: str | None = None, email: str | None = None):
    normalized_email = normalize_email(email)

    if normalized_email:
        email_report_path = get_email_last_report_path(normalized_email)
        if email_report_path.exists():
            try:
                with open(email_report_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass

    if not client_id:
        return None

    report_path = get_client_last_report_path(client_id)

    if not report_path.exists():
        return None

    try:
        with open(report_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def safe_float(value):
    try:
        if value is None or value == "":
            return 0.0

        text = str(value).strip()

        # limpiar símbolos
        text = text.replace("Bs.", "").replace("Bs", "").replace("$", "").strip()

        is_percent = "%" in text
        text = text.replace("%", "").strip()

        # formato venezolano: 1.234,56 → 1234.56
        if "," in text:
            text = text.replace(".", "").replace(",", ".")

        number = float(text)

        # 🔥 clave: evitar multiplicación por 100
        if is_percent:
            number = number / 100

        return number

    except Exception:
        return 0.0


def infer_signal(value):
    try:
        v = float(value)
    except Exception:
        return "revisar"

    if v > 0.5:
        return "subir"
    if v < -0.5:
        return "bajar"
    return "mantener"


def format_money(value):
    try:
        return f"{float(value):.2f}"
    except Exception:
        return "0.00"


def format_pct(value):
    try:
        n = float(value)
        sign = "+" if n > 0 else ""
        return f"{sign}{n:.2f}%"
    except Exception:
        return "0.00%"
    

FX_HISTORY_CACHE = {
    "data": [],
    "loaded_at": None
}

def load_fx_history(force_refresh=False):
    now = datetime.now(timezone.utc)

    if not force_refresh and FX_HISTORY_CACHE["loaded_at"]:
        age = (now - FX_HISTORY_CACHE["loaded_at"]).total_seconds()
        if age < 300:
            return FX_HISTORY_CACHE["data"]

    try:
        response = requests.get(SHEET_API, timeout=8)
        response.raise_for_status()
        data = response.json()
        data = data if isinstance(data, list) else []

        FX_HISTORY_CACHE["data"] = data
        FX_HISTORY_CACHE["loaded_at"] = now

        return data
    except Exception as e:
        print("ERROR leyendo Google Sheet:", e)
        return FX_HISTORY_CACHE["data"] or []


def save_fx_history(data):
    # Ya no usamos JSON local
    return None


def save_fx_entry(entry):
    try:
        response = requests.post(SHEET_API, json=entry, timeout=20)
        response.raise_for_status()

        text = response.text
        print("RESPONSE SHEET:", text)

        try:
            data = response.json()
        except Exception:
            return f"error: invalid json → {text[:200]}"

        return data.get("status", "unknown")

    except Exception as e:
        print("ERROR guardando en Google Sheet:", e)
        return f"error: {str(e)}"
        

def calculate_compuesto(entry):
    # Tasa Recomendada Garciabanchs (TRG):
    # Si existe mercado P2P, se calcula como 50% del mayor P2P (Binance o Bybit),
    # 30% del dólar promedio y 20% del Euro BCV.
    # Si no hay P2P disponible, se usa el dólar promedio como referencia principal
    # para evitar distorsiones por datos incompletos.
    # Si tampoco hay dólar promedio, se usa Euro BCV; si no, Dólar BCV.

    monitor = entry.get("monitor") or 0
    bcv = entry.get("bcv") or 0
    binance = entry.get("binance") or 0
    bybit = entry.get("usdt") or 0
    euro = entry.get("dolartoday") or 0

    try:
        monitor = float(monitor)
        bcv = float(bcv)
        binance = float(binance)
        bybit = float(bybit)
        euro = float(euro)
    except:
        return 0

    p2p_max = max(binance, bybit)

    if p2p_max > 0:
        compuesto = 0.5 * p2p_max

        if monitor > 0:
            compuesto += 0.3 * monitor

        if euro > 0:
            compuesto += 0.2 * euro

        return round(compuesto, 2)

    if monitor > 0:
        return round(monitor, 2)

    if euro > 0:
        return round(euro, 2)

    if bcv > 0:
        return round(bcv, 2)

    return 0

def fx_entry_exists(history, date):
    target = str(date).strip()[:10]

    for item in history:
        item_date = str(item.get("date", "")).strip()[:10]
        if item_date == target:
            return True

    return False

def normalize_fx_entry(data: dict, date: str | None = None) -> dict:
    entry = {
        "date": date or str(data.get("date") or datetime.now().strftime("%Y-%m-%d")),
        "bcv": safe_float(data.get("bcv")),
        "monitor": safe_float(data.get("monitor")),
        "binance": safe_float(data.get("binance")),
        "usdt": safe_float(data.get("usdt")),
        "dolartoday": safe_float(data.get("dolartoday")),
    }

    entry["compuesto"] = calculate_compuesto(entry)
    return entry


def upsert_fx_history_entry(history: list, entry: dict) -> tuple[list, str]:
    date = entry.get("date")
    updated = False
    clean_history = []

    for item in history:
        if item.get("date") == date:
            clean_history.append(entry)
            updated = True
        else:
            clean_history.append(item)

    if not updated:
        clean_history.append(entry)

    clean_history.sort(key=lambda x: x.get("date", ""))
    return clean_history, "updated" if updated else "saved"


def fx_label_map(key: str) -> str:
    mapping = {
        "bcv": "Dólar BCV",
        "monitor": "Dólar promedio",
        "compuesto": "Compuesto",
        "binance": "Binance P2P",
        "usdt": "Bybit P2P",
        "dolartoday": "Euro BCV"
    }
    return mapping.get(key, key)


def report_type_label(report_type: str) -> str:
    mapping = {
        "operativo": "Reporte operativo de actualización de precios",
        "ejecutivo": "Reporte ejecutivo de actualización de precios",
        "tienda": "Reporte de tienda para cambio de precios",
    }
    return mapping.get(report_type, "Reporte de actualización de precios")


def validate_report_type(report_type: str) -> str:
    if report_type not in VALID_REPORT_TYPES:
        raise HTTPException(status_code=404, detail=f"No existe el tipo de reporte {report_type}.")
    return report_type


def validate_fx_key(fx_key: str) -> str:
    if fx_key not in VALID_FX_KEYS:
        raise HTTPException(status_code=404, detail=f"No existe la referencia {fx_key}.")
    return fx_key


def validate_html_view_type(view_type: str) -> str:
    if view_type not in {"desktop", "mobile"}:
        raise HTTPException(status_code=404, detail=f"No existe la vista HTML {view_type}.")
    return view_type


def build_market_downloads(market_key: str) -> dict:
    pdfs = {}
    for report_type in ["operativo", "ejecutivo", "tienda"]:
        pdfs[report_type] = {}
        for fx_key in ["compuesto", "monitor", "binance", "usdt", "dolartoday"]:
            pdfs[report_type][fx_key] = f"/report/pdf/{market_key}/{report_type}/{fx_key}"

    return {
        "pdfs": pdfs,
        "zip": "#"
    }


def sanitize_filename(text: str) -> str:
    safe = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in text.strip())
    while "__" in safe:
        safe = safe.replace("__", "_")
    return safe.strip("_") or "archivo"


def safe_text(value, default="—"):
    text = str(value).strip() if value is not None else ""
    return text if text else default


def build_pdf_styles():
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name="HeroTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=25,
        leading=27,
        textColor=colors.HexColor("#064e3b"),
        alignment=TA_LEFT,
        spaceAfter=4,
    ))

    styles.add(ParagraphStyle(
        name="HeroSub",
        parent=styles["Heading2"],
        fontName="Helvetica",
        fontSize=12.2,
        leading=14.5,
        textColor=colors.HexColor("#334155"),
        alignment=TA_LEFT,
        spaceAfter=8,
    ))

    styles.add(ParagraphStyle(
        name="SectionTitleGreen",
        parent=styles["Heading3"],
        fontName="Helvetica-Bold",
        fontSize=14.2,
        leading=16.5,
        textColor=colors.HexColor("#064e3b"),
        spaceAfter=6,
    ))

    styles.add(ParagraphStyle(
        name="SectionTitleGreenCenter",
        parent=styles["SectionTitleGreen"],
        alignment=TA_CENTER,
    ))

    styles.add(ParagraphStyle(
        name="LabelSmall",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.0,
        leading=9.2,
        textColor=colors.HexColor("#64748b"),
    ))

    styles.add(ParagraphStyle(
        name="ValueBig",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=17.5,
        leading=18.5,
        textColor=colors.HexColor("#0f172a"),
    ))

    styles.add(ParagraphStyle(
        name="CardText",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9.2,
        leading=11.7,
        textColor=colors.HexColor("#334155"),
    ))

    styles.add(ParagraphStyle(
        name="BodySoft",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10.2,
        leading=13.4,
        textColor=colors.HexColor("#334155"),
    ))

    styles.add(ParagraphStyle(
        name="SmallMuted",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=11.4,
        textColor=colors.HexColor("#64748b"),
    ))

    styles.add(ParagraphStyle(
        name="CTA",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10.8,
        leading=13.6,
        textColor=colors.HexColor("#064e3b"),
    ))

    styles.add(ParagraphStyle(
        name="HeroAuthorLabel",
        parent=styles["SmallMuted"],
        fontName="Helvetica",
        fontSize=9.2,
        leading=11.5,
        textColor=colors.HexColor("#64748b"),
        alignment=TA_CENTER,
        spaceAfter=2,
    ))

    styles.add(ParagraphStyle(
        name="HeroAuthorName",
        parent=styles["HeroSub"],
        fontName="Helvetica-Bold",
        fontSize=12.2,
        leading=13.5,
        textColor=colors.HexColor("#064e3b"),
        alignment=TA_CENTER,
        spaceAfter=0,
    ))

    styles.add(ParagraphStyle(
        name="TableHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.2,
        leading=9.4,
        textColor=colors.white,
        alignment=TA_CENTER,
    ))

    styles.add(ParagraphStyle(
        name="TableCell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.4,
        leading=9.0,
        textColor=colors.HexColor("#334155"),
        alignment=TA_LEFT,
    ))

    styles.add(ParagraphStyle(
        name="TableCellBold",
        parent=styles["TableCell"],
        fontName="Helvetica-Bold",
    ))

    styles.add(ParagraphStyle(
        name="TableCellRight",
        parent=styles["TableCell"],
        alignment=TA_RIGHT,
    ))

    styles.add(ParagraphStyle(
        name="TableCellRightBold",
        parent=styles["TableCellBold"],
        alignment=TA_RIGHT,
    ))

    styles.add(ParagraphStyle(
        name="TableCellCenter",
        parent=styles["TableCell"],
        alignment=TA_CENTER,
    ))

    return styles


def build_info_table(company_name, report_date, market_label, city, fx_label, fx_value, fx_var, report_type="operativo"):
    if report_type == "tienda":
        reference_value = "No aplica"
        fx_value_text = "No aplica"
        fx_var_text = "No aplica"
        motor_value = "PrecioDólar LIVE"
    else:
        reference_value = safe_text(fx_label)
        fx_value_text = f"Bs. {format_money(fx_value)}"
        fx_var_text = format_pct(fx_var)
        motor_value = "PrecioDólar LIVE"

    inner_data = [
        [
            Paragraph("<b>Empresa</b>", PDF_STYLES["LabelSmall"]),
            Paragraph("<b>Fecha</b>", PDF_STYLES["LabelSmall"]),
            Paragraph("<b>Mercado</b>", PDF_STYLES["LabelSmall"]),
            Paragraph("<b>Ubicación</b>", PDF_STYLES["LabelSmall"]),
        ],
        [
            Paragraph(safe_text(company_name), PDF_STYLES["CardText"]),
            Paragraph(safe_text(report_date), PDF_STYLES["CardText"]),
            Paragraph(safe_text(market_label), PDF_STYLES["CardText"]),
            Paragraph(safe_text(city), PDF_STYLES["CardText"]),
        ],
        [
            Paragraph("<b>Referencia</b>", PDF_STYLES["LabelSmall"]),
            Paragraph("<b>Tipo de cambio</b>", PDF_STYLES["LabelSmall"]),
            Paragraph("<b>Variación</b>", PDF_STYLES["LabelSmall"]),
            Paragraph("<b>Motor</b>", PDF_STYLES["LabelSmall"]),
        ],
        [
            Paragraph(reference_value, PDF_STYLES["CardText"]),
            Paragraph(fx_value_text, PDF_STYLES["CardText"]),
            Paragraph(fx_var_text, PDF_STYLES["CardText"]),
            Paragraph(motor_value, PDF_STYLES["CardText"]),
        ],
    ]

    inner_table = Table(inner_data, colWidths=[46 * mm, 35 * mm, 51 * mm, 43 * mm])
    inner_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#e2e8f0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    outer = Table([[inner_table]], colWidths=[175 * mm])
    outer.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    return outer


def build_kpi_cards(report_data, fx_key, market_rows_count, report_type="operativo"):
    fx_boxes = report_data.get("fx_boxes", {})
    fx_box = fx_boxes.get(fx_key, {})

    if report_type == "tienda":
        kpi_data = [[
            Paragraph("Productos válidos", PDF_STYLES["LabelSmall"]),
            Paragraph("Referencia activa", PDF_STYLES["LabelSmall"]),
            Paragraph("Variación referencia", PDF_STYLES["LabelSmall"]),
        ], [
            Paragraph(str(market_rows_count), PDF_STYLES["ValueBig"]),
            Paragraph("No aplica", PDF_STYLES["ValueBig"]),
            Paragraph("No aplica", PDF_STYLES["ValueBig"]),
        ]]
    else:
        kpi_data = [[
            Paragraph("Productos válidos", PDF_STYLES["LabelSmall"]),
            Paragraph("Referencia activa", PDF_STYLES["LabelSmall"]),
            Paragraph("Variación referencia", PDF_STYLES["LabelSmall"]),
        ], [
            Paragraph(str(market_rows_count), PDF_STYLES["ValueBig"]),
            Paragraph(fx_label_map(fx_key), PDF_STYLES["ValueBig"]),
            Paragraph(format_pct(fx_box.get("var", 0)), PDF_STYLES["ValueBig"]),
        ]]

    table = Table(kpi_data, colWidths=[58 * mm, 58 * mm, 59 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#ecfdf5")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#bbf7d0")),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dcfce7")),
        ("LEFTPADDING", (0, 0), (-1, -1), 9),
        ("RIGHTPADDING", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    return table


def build_author_books_contact_block():
    from reportlab.graphics.barcode import qr
    from reportlab.graphics.shapes import Drawing

    author_photo_path = ASSETS_DIR / "imagen-circular.png"
    book1_path = ASSETS_DIR / "camino-a-la-riqueza.jpg"
    book2_path = ASSETS_DIR / "the-artificial-intelligence-millionaire.jpg"

    amazon_author_link = "https://www.amazon.com/author/garciabanchs"
    book1_link = "https://www.amazon.com/dp/B0C9SK1MMQ"
    book2_link = "https://www.amazon.com/dp/B0D56VRYG1"
    linktree_link = "https://linktr.ee/garciabanchs"

    def make_qr(url, size_mm=28):
        qr_code = qr.QrCodeWidget(url)
        bounds = qr_code.getBounds()
        width = bounds[2] - bounds[0]
        height = bounds[3] - bounds[1]
        d = Drawing(
            size_mm * mm,
            size_mm * mm,
            transform=[size_mm * mm / width, 0, 0, size_mm * mm / height, 0, 0]
        )
        d.add(qr_code)
        return d

    author_block = [Paragraph("Acerca del autor", PDF_STYLES["SectionTitleGreen"]), Spacer(1, 4)]

    if author_photo_path.exists():
        author_img = Image(str(author_photo_path), width=28 * mm, height=28 * mm)
        author_header = Table([[
            author_img,
            Paragraph("<b>Ángel García Banchs</b>", PDF_STYLES["HeroSub"])
        ]], colWidths=[31 * mm, 49 * mm])
        author_header.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        author_block.append(author_header)
        author_block.append(Spacer(1, 8))
    else:
        author_block.append(Paragraph("<b>Ángel García Banchs</b>", PDF_STYLES["HeroSub"]))
        author_block.append(Spacer(1, 6))

    author_block.append(Paragraph(
        "Economista, académico, consultor, mentor y asesor. Especializado en análisis económico, finanzas, pricing, automatización y herramientas de alto valor para empresas.",
        PDF_STYLES["BodySoft"]
    ))

    author_inner = Table(
        [
            [author_block],
            [Spacer(1, 30 * mm)]
        ],
        colWidths=[80 * mm]
    )
    author_inner.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    author_card = Table([[author_inner]], colWidths=[80 * mm], rowHeights=[118 * mm])
    author_card.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#cbd5e1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    if book1_path.exists():
        book1_img = Image(str(book1_path), width=18 * mm, height=26 * mm)
    else:
        book1_img = Paragraph("Sin portada", PDF_STYLES["SmallMuted"])

    if book2_path.exists():
        book2_img = Image(str(book2_path), width=18 * mm, height=26 * mm)
    else:
        book2_img = Paragraph("Sin portada", PDF_STYLES["SmallMuted"])

    books_rows = [
        [
            book1_img,
            Paragraph(
                f'<link href="{book1_link}"><b>Camino a la Riqueza</b></link><br/><link href="{book1_link}">Ver en Amazon</link>',
                PDF_STYLES["BodySoft"]
            )
        ],
        [
            book2_img,
            Paragraph(
                f'<link href="{book2_link}"><b>The Artificial Intelligence Millionaire</b></link><br/><link href="{book2_link}">Ver en Amazon</link>',
                PDF_STYLES["BodySoft"]
            )
        ],
    ]

    books_table = Table(books_rows, colWidths=[22 * mm, 48 * mm])
    books_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#e2e8f0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LINK", (0, 0), (0, 0), book1_link),
        ("LINK", (1, 0), (1, 0), book1_link),
        ("LINK", (0, 1), (0, 1), book2_link),
        ("LINK", (1, 1), (1, 1), book2_link),
    ]))

    books_inner = Table(
        [
            [books_table],
            [Spacer(1, 8 * mm)],
            [make_qr(amazon_author_link, size_mm=28)]
        ],
        colWidths=[70 * mm],
    )
    books_inner.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    books_block = [
        Paragraph("Libros", PDF_STYLES["SectionTitleGreen"]),
        Spacer(1, 4),
        books_inner
    ]

    books_card = Table([[books_block]], colWidths=[80 * mm], rowHeights=[118 * mm])
    books_card.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#cbd5e1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    contact_top = [
        Paragraph("Contacto", PDF_STYLES["SectionTitleGreen"]),
        Spacer(1, 4),
        Paragraph(
            "Accede a todos mis contenidos, asesorías y redes:",
            PDF_STYLES["BodySoft"]
        ),
        Spacer(1, 6),
        Paragraph(
            f'<link href="{linktree_link}"><b>Linktree: garciabanchs</b></link>',
            PDF_STYLES["CTA"]
        ),
    ]

    contact_inner = Table(
        [
            [contact_top],
            [Spacer(1, 8 * mm)],
            [make_qr(linktree_link, size_mm=28)]
        ],
        colWidths=[75 * mm],
    )
    contact_inner.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    contact_fill = Table(
        [
            [contact_inner],
            [Spacer(1, 22 * mm)]
        ],
        colWidths=[85 * mm]
    )
    contact_fill.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    contact_card = Table([[contact_fill]], colWidths=[85 * mm], rowHeights=[118 * mm])
    contact_card.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#cbd5e1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    container = Table(
        [[
            author_card,
            "",
            books_card,
            "",
            contact_card
        ]],
        colWidths=[82 * mm, 10 * mm, 82 * mm, 10 * mm, 81 * mm],
        rowHeights=[118 * mm]
    )
    container.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    return container


PDF_STYLES = build_pdf_styles()


def get_report_data_or_raise(request: Request):
    client_id = get_client_id_from_request(request)
    email = get_email_from_request(request)
    if email:
        report_data = load_last_report(None, email=email)
    else:
        report_data = load_last_report(client_id)

    if report_data is None:
        raise HTTPException(status_code=404, detail="No hay reporte persistido todavía para este cliente o email.")

    return report_data


def find_market(report_data: dict, market_key: str):
    for market in report_data.get("markets", []):
        if market.get("market_key") == market_key:
            return market
    return None


def format_report_date(value):
    from datetime import datetime

    if not value:
        return "—"

    try:
        dt = datetime.fromisoformat(str(value))
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return str(value)


def build_executive_or_operational_note(report_type: str):
    if report_type == "tienda":
        title = "Recordatorio para uso en tienda"
        text = (
            "Este material puede utilizarse en PDF o impreso para actualizar precios en el establecimiento. "
            "También existe una versión HTML descargable para escritorio y una versión móvil con tarjetas, pensada para recorrer la tienda, farmacia o supermercado e ir cambiando precios producto por producto con mayor comodidad."
        )
    elif report_type == "ejecutivo":
        title = "Lectura ejecutiva"
        text = (
            "Este reporte resume los precios sugeridos para apoyar decisiones gerenciales. "
            "La tabla ejecutiva prioriza producto, precio anterior, precio sugerido y variación, eliminando columnas operativas de comparación competitiva para facilitar una revisión rápida."
        )
    else:
        title = "Lectura operativa"
        text = (
            "Este reporte presenta una actualización de precios basada en la evolución reciente del tipo de cambio y su impacto sobre la reposición de inventario. "
            "El objetivo no es solo ajustar precios, sino preservar capital, competitividad y capacidad operativa.<br/><br/>"
            "Los precios sugeridos integran dos dimensiones clave: presión competitiva y presión cambiaria. "
            "La combinación de ambas permite decisiones más robustas que el simple seguimiento de una sola referencia."
        )

    block = Table([[
        Paragraph(f"<b>{title}</b><br/><br/>{text}", PDF_STYLES["BodySoft"])
    ]], colWidths=[265 * mm])

    block.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#ecfdf5")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#bbf7d0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))

    return block


def build_prices_table(rows, report_type: str):
    if report_type == "tienda":
        table_data = [[
            Paragraph("Producto", PDF_STYLES["TableHeader"]),
            Paragraph("SKU", PDF_STYLES["TableHeader"]),
            Paragraph("Unidad", PDF_STYLES["TableHeader"]),
            Paragraph("Precio viejo", PDF_STYLES["TableHeader"]),
            Paragraph("Precio nuevo", PDF_STYLES["TableHeader"]),
            Paragraph("Señal", PDF_STYLES["TableHeader"]),
        ]]

        for row in rows:
            table_data.append([
                Paragraph(safe_text(row.get("nombre_producto", "")), PDF_STYLES["TableCellBold"]),
                Paragraph(safe_text(row.get("sku", "")), PDF_STYLES["TableCell"]),
                Paragraph(safe_text(row.get("unidad", "")), PDF_STYLES["TableCell"]),
                Paragraph(format_money(row.get("precio_viejo_usd", 0)), PDF_STYLES["TableCellRight"]),
                Paragraph(format_money(row.get("precio_nuevo_usd", 0)), PDF_STYLES["TableCellRightBold"]),
                Paragraph(str(row.get("senal", "")).capitalize(), PDF_STYLES["TableCellCenter"]),
            ])

        col_widths = [85 * mm, 42 * mm, 32 * mm, 30 * mm, 30 * mm, 30 * mm]

    elif report_type == "ejecutivo":
        table_data = [[
            Paragraph("Producto", PDF_STYLES["TableHeader"]),
            Paragraph("SKU", PDF_STYLES["TableHeader"]),
            Paragraph("Unidad", PDF_STYLES["TableHeader"]),
            Paragraph("Precio viejo", PDF_STYLES["TableHeader"]),
            Paragraph("Precio nuevo", PDF_STYLES["TableHeader"]),
            Paragraph("%", PDF_STYLES["TableHeader"]),
        ]]

        for row in rows:
            table_data.append([
                Paragraph(safe_text(row.get("nombre_producto", "")), PDF_STYLES["TableCellBold"]),
                Paragraph(safe_text(row.get("sku", "")), PDF_STYLES["TableCell"]),
                Paragraph(safe_text(row.get("unidad", "")), PDF_STYLES["TableCell"]),
                Paragraph(format_money(row.get("precio_viejo_usd", 0)), PDF_STYLES["TableCellRight"]),
                Paragraph(format_money(row.get("precio_nuevo_usd", 0)), PDF_STYLES["TableCellRightBold"]),
                Paragraph(format_pct(row.get("cambio_pct", 0)), PDF_STYLES["TableCellRight"]),
            ])

        col_widths = [90 * mm, 45 * mm, 35 * mm, 30 * mm, 30 * mm, 25 * mm]

    else:
        table_data = [[
            Paragraph("Producto", PDF_STYLES["TableHeader"]),
            Paragraph("SKU", PDF_STYLES["TableHeader"]),
            Paragraph("Unidad", PDF_STYLES["TableHeader"]),
            Paragraph("Precio viejo", PDF_STYLES["TableHeader"]),
            Paragraph("Precio nuevo", PDF_STYLES["TableHeader"]),
            Paragraph("%", PDF_STYLES["TableHeader"]),
            Paragraph("Líder", PDF_STYLES["TableHeader"]),
            Paragraph("Intermedio", PDF_STYLES["TableHeader"]),
            Paragraph("Económico", PDF_STYLES["TableHeader"]),
            Paragraph("Peso C", PDF_STYLES["TableHeader"]),
            Paragraph("Peso R", PDF_STYLES["TableHeader"]),
            Paragraph("Señal", PDF_STYLES["TableHeader"]),
        ]]

        for row in rows:
            table_data.append([
                Paragraph(safe_text(row.get("nombre_producto", "")), PDF_STYLES["TableCellBold"]),
                Paragraph(safe_text(row.get("sku", "")), PDF_STYLES["TableCell"]),
                Paragraph(safe_text(row.get("unidad", "")), PDF_STYLES["TableCell"]),
                Paragraph(format_money(row.get("precio_viejo_usd", 0)), PDF_STYLES["TableCellRight"]),
                Paragraph(format_money(row.get("precio_nuevo_usd", 0)), PDF_STYLES["TableCellRightBold"]),
                Paragraph(format_pct(row.get("cambio_pct", 0)), PDF_STYLES["TableCellRight"]),
                Paragraph(format_money(row.get("competidor_lider", 0)), PDF_STYLES["TableCellRight"]),
                Paragraph(format_money(row.get("competidor_intermedio", 0)), PDF_STYLES["TableCellRight"]),
                Paragraph(format_money(row.get("competidor_economico", 0)), PDF_STYLES["TableCellRight"]),
                Paragraph(f"{safe_float(row.get('peso_competencia', 0)) * 100:.0f}%", PDF_STYLES["TableCellCenter"]),
                Paragraph(f"{safe_float(row.get('peso_riesgo', 0)) * 100:.0f}%", PDF_STYLES["TableCellCenter"]),
                Paragraph(str(row.get("senal", "")).capitalize(), PDF_STYLES["TableCellCenter"]),
            ])

        col_widths = [
            49 * mm, 24 * mm, 19 * mm, 18 * mm, 18 * mm, 13 * mm,
            16 * mm, 19 * mm, 19 * mm, 12 * mm, 12 * mm, 15 * mm
        ]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#064e3b")),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))

    return table


def apply_excel_header_style(ws, row_number: int):
    header_fill = PatternFill("solid", fgColor="064E3B")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")

    for cell in ws[row_number]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_excel_table_style(ws):
    thin = Side(style="thin", color="CBD5E1")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def autosize_columns(ws, min_width=12, max_width=42):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, min_width), max_width)


def generate_market_fx_excel(report_data: dict, market_key: str, fx_key: str) -> Path:
    ensure_data_dir()
    validate_fx_key(fx_key)

    market = find_market(report_data, market_key)
    if not market:
        raise HTTPException(status_code=404, detail=f"No existe el mercado {market_key}.")

    fx_views = market.get("fx_views", {})
    fx_data = fx_views.get(fx_key)
    if not fx_data:
        raise HTTPException(status_code=404, detail=f"No existe la referencia {fx_key} para {market_key}.")

    rows = fx_data.get("rows", [])

    company_name = report_data.get("company_name", "Empresa")
    report_date = format_report_date(report_data.get("report_date", ""))
    market_label = market.get("market_label", market_key)
    city = market.get("city", "")
    fx_label = fx_label_map(fx_key)

    fx_boxes = report_data.get("fx_boxes", {})
    fx_box = fx_boxes.get(fx_key, {})
    fx_value = fx_box.get("valor", 0)
    fx_var = fx_box.get("var", 0)

    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_tabla = wb.create_sheet("Tabla operativa")
    ws_autor = wb.create_sheet("Autor y contacto")

    title_fill = PatternFill("solid", fgColor="ECFDF5")
    title_font = Font(color="064E3B", bold=True, size=14)

    ws_resumen["A1"] = "PrecioDólar LIVE · Excel operativo"
    ws_resumen["A1"].font = Font(color="064E3B", bold=True, size=16)
    ws_resumen["A1"].fill = title_fill

    resumen_data = [
        ("Empresa", safe_text(company_name)),
        ("Fecha", safe_text(report_date)),
        ("Mercado", safe_text(market_label)),
        ("Ubicación", safe_text(city)),
        ("Referencia activa", safe_text(fx_label)),
        ("Tipo de cambio", f"Bs. {format_money(fx_value)}"),
        ("Variación referencia", format_pct(fx_var)),
        ("Productos válidos", len(rows)),
        ("Motor", "PrecioDólar LIVE"),
        (
            "Lectura operativa",
            "Este Excel operativo presenta precios sugeridos a partir de la presión competitiva y la presión cambiaria. "
            "La finalidad es facilitar una actualización ordenada de precios, preservando capital, competitividad y capacidad de reposición."
        ),
    ]

    ws_resumen.append([])
    for label, value in resumen_data:
        ws_resumen.append([label, value])

    for row in ws_resumen.iter_rows(min_row=3, max_row=12, min_col=1, max_col=2):
        row[0].font = Font(bold=True, color="064E3B")
        row[0].fill = PatternFill("solid", fgColor="F8FAFC")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    ws_resumen.merge_cells("A1:B1")
    ws_resumen["A1"].alignment = Alignment(horizontal="center")
    apply_excel_table_style(ws_resumen)
    autosize_columns(ws_resumen, min_width=18, max_width=70)

    headers = [
        "Producto",
        "SKU",
        "Unidad",
        "Precio viejo USD",
        "Precio nuevo USD",
        "Cambio %",
        "Competidor líder",
        "Competidor intermedio",
        "Competidor económico",
        "Peso competencia",
        "Peso riesgo cambiario",
        "Señal",
    ]
    ws_tabla.append(headers)
    apply_excel_header_style(ws_tabla, 1)

    for row in rows:
        ws_tabla.append([
            safe_text(row.get("nombre_producto", "")),
            safe_text(row.get("sku", "")),
            safe_text(row.get("unidad", "")),
            safe_float(row.get("precio_viejo_usd", 0)),
            safe_float(row.get("precio_nuevo_usd", 0)),
            safe_float(row.get("cambio_pct", 0)) / 100,
            safe_float(row.get("competidor_lider", 0)),
            safe_float(row.get("competidor_intermedio", 0)),
            safe_float(row.get("competidor_economico", 0)),
            safe_float(row.get("peso_competencia", 0)),
            safe_float(row.get("peso_riesgo", 0)),
            str(row.get("senal", "")).capitalize(),
        ])

    for data_row in ws_tabla.iter_rows(min_row=2):
        for cell in data_row[3:5]:
            cell.number_format = '$#,##0.00'
        data_row[5].number_format = '0.00%'
        for cell in data_row[6:9]:
            cell.number_format = '$#,##0.00'
        data_row[9].number_format = '0%'
        data_row[10].number_format = '0%'

    ws_tabla.freeze_panes = "A2"
    ws_tabla.auto_filter.ref = ws_tabla.dimensions
    apply_excel_table_style(ws_tabla)
    autosize_columns(ws_tabla, min_width=12, max_width=38)

    ws_autor["A1"] = "Autor y contacto"
    ws_autor["A1"].font = title_font
    ws_autor["A1"].fill = title_fill

    autor_data = [
        ("Nombre", "Ángel García Banchs"),
        (
            "Perfil profesional",
            "Economista, académico, consultor, mentor y asesor. Especializado en análisis económico, finanzas, pricing, automatización y herramientas de alto valor para empresas."
        ),
        ("WhatsApp", "https://api.whatsapp.com/send?phone=34622197658"),
        ("Amazon autor", "https://www.amazon.com/author/garciabanchs"),
        ("Libros", "Camino a la Riqueza; The Artificial Intelligence Millionaire"),
        ("Linktree", "https://linktr.ee/garciabanchs"),
    ]

    ws_autor.append([])
    for label, value in autor_data:
        ws_autor.append([label, value])

    for row in ws_autor.iter_rows(min_row=3, max_row=8, min_col=1, max_col=2):
        row[0].font = Font(bold=True, color="064E3B")
        row[0].fill = PatternFill("solid", fgColor="F8FAFC")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    ws_autor.merge_cells("A1:B1")
    ws_autor["A1"].alignment = Alignment(horizontal="center")
    apply_excel_table_style(ws_autor)
    autosize_columns(ws_autor, min_width=18, max_width=75)

    for ws in [ws_resumen, ws_tabla, ws_autor]:
        ws.sheet_view.showGridLines = False

    base_filename = sanitize_filename(f"{company_name}_{market_key}_operativo_{fx_key}")
    excel_path = EXCEL_DIR / f"{base_filename}.xlsx"
    wb.save(excel_path)

    return excel_path


def generate_market_fx_html(report_data: dict, market_key: str, view_type: str, fx_key: str) -> Path:
    ensure_data_dir()
    validate_html_view_type(view_type)
    validate_fx_key(fx_key)

    market = find_market(report_data, market_key)
    if not market:
        raise HTTPException(status_code=404, detail=f"No existe el mercado {market_key}.")

    fx_views = market.get("fx_views", {})
    fx_data = fx_views.get(fx_key)
    if not fx_data:
        raise HTTPException(status_code=404, detail=f"No existe la referencia {fx_key} para {market_key}.")

    rows = fx_data.get("rows", [])

    company_name = report_data.get("company_name", "Empresa")
    report_date = format_report_date(report_data.get("report_date", ""))
    market_label = market.get("market_label", market_key)
    city = market.get("city", "")
    fx_label = fx_label_map(fx_key)

    fx_boxes = report_data.get("fx_boxes", {})
    fx_box = fx_boxes.get(fx_key, {})
    fx_value = fx_box.get("valor", 0)
    fx_var = fx_box.get("var", 0)

    title = f"PrecioDólar LIVE · HTML {'Desktop' if view_type == 'desktop' else 'Móvil'}"

    css = """
    body {
        font-family: Arial, Helvetica, sans-serif;
        margin: 24px;
        color: #0f172a;
        background: #f8fafc;
    }
    .container {
        max-width: 1200px;
        margin: 0 auto;
        background: #ffffff;
        border: 1px solid #cbd5e1;
        border-radius: 12px;
        padding: 22px;
    }
    h1 {
        color: #064e3b;
        margin-bottom: 4px;
    }
    h2 {
        color: #064e3b;
        margin-top: 24px;
    }
    .muted {
        color: #64748b;
        font-size: 14px;
    }
    .summary {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
        gap: 10px;
        margin-top: 18px;
    }
    .summary-item {
        background: #ecfdf5;
        border: 1px solid #bbf7d0;
        border-radius: 10px;
        padding: 12px;
    }
    .label {
        font-size: 12px;
        color: #64748b;
        font-weight: bold;
        text-transform: uppercase;
    }
    .value {
        margin-top: 4px;
        font-size: 16px;
        font-weight: bold;
    }
    table {
        width: 100%;
        border-collapse: collapse;
        margin-top: 16px;
        background: #ffffff;
    }
    th {
        background: #064e3b;
        color: #ffffff;
        padding: 10px;
        text-align: left;
        font-size: 13px;
    }
    td {
        border: 1px solid #e2e8f0;
        padding: 9px;
        font-size: 13px;
        vertical-align: top;
    }
    tr:nth-child(even) {
        background: #f8fafc;
    }
    .card {
        border: 1px solid #cbd5e1;
        border-radius: 12px;
        padding: 14px;
        margin: 12px 0;
        background: #ffffff;
    }
    .card-title {
        font-weight: bold;
        color: #064e3b;
        font-size: 17px;
        margin-bottom: 8px;
    }
    .field {
        margin: 5px 0;
    }
    .signal {
        font-weight: bold;
        color: #064e3b;
    }
    .footer {
        margin-top: 28px;
        color: #64748b;
        font-size: 13px;
    }
    """

    summary_html = f"""
    <div class="summary">
        <div class="summary-item"><div class="label">Empresa</div><div class="value">{escape(safe_text(company_name))}</div></div>
        <div class="summary-item"><div class="label">Fecha</div><div class="value">{escape(safe_text(report_date))}</div></div>
        <div class="summary-item"><div class="label">Mercado</div><div class="value">{escape(safe_text(market_label))}</div></div>
        <div class="summary-item"><div class="label">Ubicación</div><div class="value">{escape(safe_text(city))}</div></div>
        <div class="summary-item"><div class="label">Referencia activa</div><div class="value">{escape(safe_text(fx_label))}</div></div>
        <div class="summary-item"><div class="label">Tipo de cambio</div><div class="value">Bs. {escape(format_money(fx_value))}</div></div>
        <div class="summary-item"><div class="label">Variación referencia</div><div class="value">{escape(format_pct(fx_var))}</div></div>
        <div class="summary-item"><div class="label">Productos válidos</div><div class="value">{len(rows)}</div></div>
    </div>
    """

    if view_type == "desktop":
        rows_html = ""
        for row in rows:
            rows_html += f"""
            <tr>
                <td>{escape(safe_text(row.get("nombre_producto", "")))}</td>
                <td>{escape(safe_text(row.get("sku", "")))}</td>
                <td>{escape(safe_text(row.get("unidad", "")))}</td>
                <td>{escape(format_money(row.get("precio_viejo_usd", 0)))}</td>
                <td><b>{escape(format_money(row.get("precio_nuevo_usd", 0)))}</b></td>
                <td>{escape(format_pct(row.get("cambio_pct", 0)))}</td>
                <td class="signal">{escape(str(row.get("senal", "")).capitalize())}</td>
            </tr>
            """

        body_content = f"""
        <h2>Tabla operativa</h2>
        <table>
            <thead>
                <tr>
                    <th>Producto</th>
                    <th>SKU</th>
                    <th>Unidad</th>
                    <th>Precio viejo USD</th>
                    <th>Precio nuevo USD</th>
                    <th>Cambio %</th>
                    <th>Señal</th>
                </tr>
            </thead>
            <tbody>
                {rows_html if rows_html else '<tr><td colspan="8">Este mercado no tiene productos válidos.</td></tr>'}
            </tbody>
        </table>
        """
    else:
        cards_html = ""
        for row in rows:
            cards_html += f"""
            <div class="card">
                <div class="card-title">{escape(safe_text(row.get("nombre_producto", "")))}</div>
                <div class="field"><b>SKU:</b> {escape(safe_text(row.get("sku", "")))}</div>
                <div class="field"><b>Unidad:</b> {escape(safe_text(row.get("unidad", "")))}</div>
                <div class="field"><b>Precio viejo USD:</b> {escape(format_money(row.get("precio_viejo_usd", 0)))}</div>
                <div class="field"><b>Precio nuevo USD:</b> {escape(format_money(row.get("precio_nuevo_usd", 0)))}</div>
                <div class="field"><b>Cambio %:</b> {escape(format_pct(row.get("cambio_pct", 0)))}</div>
                <div class="field signal"><b>Señal:</b> {escape(str(row.get("senal", "")).capitalize())}</div>
            </div>
            """

        body_content = f"""
        <h2>Tarjetas móviles</h2>
        {cards_html if cards_html else '<div class="card">Este mercado no tiene productos válidos.</div>'}
        """

    html_content = f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="utf-8">
    <title>{escape(title)}</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>{css}</style>
</head>
<body>
    <div class="container">
        <h1>{escape(title)}</h1>
        <div class="muted">Documento generado automáticamente por PrecioDólar LIVE.</div>

        {summary_html}

        <h2>Lectura operativa</h2>
        <p>
            Este HTML operativo presenta precios sugeridos para facilitar la actualización de precios,
            preservando capital, competitividad y capacidad de reposición.
        </p>

        {body_content}

        <div class="footer">
            Herramienta creada por Ángel García Banchs ·
            <a href="https://linktr.ee/garciabanchs">Linktree</a> ·
            <a href="https://www.amazon.com/author/garciabanchs">Amazon autor</a>
        </div>
    </div>
</body>
</html>
"""

    base_filename = sanitize_filename(f"{company_name}_{market_key}_{view_type}_{fx_key}")
    html_path = HTML_DIR / f"{base_filename}.html"

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    return html_path


def generate_market_fx_pdf(report_data: dict, market_key: str, report_type: str, fx_key: str) -> Path:
    ensure_data_dir()

    validate_report_type(report_type)
    validate_fx_key(fx_key)

    market = find_market(report_data, market_key)
    if not market:
        raise HTTPException(status_code=404, detail=f"No existe el mercado {market_key}.")

    fx_views = market.get("fx_views", {})
    fx_data = fx_views.get(fx_key)
    if not fx_data:
        raise HTTPException(status_code=404, detail=f"No existe la referencia {fx_key} para {market_key}.")

    rows = fx_data.get("rows", [])

    company_name = report_data.get("company_name", "Empresa")
    report_date = report_data.get("report_date", "")
    market_label = market.get("market_label", market_key)
    city = market.get("city", "")
    fx_label = fx_label_map(fx_key)

    fx_boxes = report_data.get("fx_boxes", {})
    fx_box = fx_boxes.get(fx_key, {})
    fx_value = fx_box.get("valor", 0)
    fx_var = fx_box.get("var", 0)

    pdf_variant = report_type

    base_filename = sanitize_filename(f"{company_name}_{market_key}_{pdf_variant}_{fx_key}")
    filename = f"{base_filename}.pdf"
    pdf_path = PDF_DIR / filename

    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=landscape(A4),
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm
    )

    story = []

    author_photo_path = ASSETS_DIR / "imagen-circular.png"

    left_hero = []
    left_hero.append(Paragraph("PrecioDólar LIVE", PDF_STYLES["HeroTitle"]))
    left_hero.append(Paragraph(report_type_label(report_type), PDF_STYLES["HeroSub"]))
    left_hero.append(Spacer(1, 6))

    left_hero.append(build_info_table(
        company_name=company_name,
        report_date=format_report_date(report_date),
        market_label=market_label,
        city=city,
        fx_label=fx_label,
        fx_value=fx_value,
        fx_var=fx_var,
        report_type=report_type
    ))

    left_hero.append(Spacer(1, 8))

    left_hero.append(build_kpi_cards(
        report_data=report_data,
        fx_key=fx_key,
        market_rows_count=len(rows),
        report_type=report_type
    ))

    right_top_cells = [""]

    if author_photo_path.exists():
        author_img = Image(str(author_photo_path), width=40 * mm, height=40 * mm)
        photo_card = Table([[author_img]], colWidths=[50 * mm])
        photo_card.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#ecfdf5")),
            ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#bbf7d0")),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        right_top_cells = [photo_card]

    right_bottom = [
        Paragraph("Herramienta creada por", PDF_STYLES["HeroAuthorLabel"]),
        Paragraph("Ángel García Banchs", PDF_STYLES["HeroAuthorName"])
    ]

    right_inner = Table([
        [right_top_cells[0]],
        [right_bottom]
    ], colWidths=[50 * mm], rowHeights=[58 * mm, 18 * mm])

    right_inner.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))

    left_card = Table([[left_hero]], colWidths=[185 * mm])
    left_card.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#cbd5e1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    right_card = Table([[right_inner]], colWidths=[70 * mm], rowHeights=[90 * mm])
    right_card.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#cbd5e1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))

    hero_table = Table([[
        left_card,
        "",
        right_card
    ]], colWidths=[185 * mm, 10 * mm, 70 * mm])

    hero_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    story.append(hero_table)
    story.append(Spacer(1, 12))
    story.append(build_executive_or_operational_note(report_type))

    story.append(PageBreak())

    if report_type == "ejecutivo":
        table_title = "Tabla ejecutiva de precios sugeridos"
    elif report_type == "tienda":
        table_title = "Tabla de tienda para cambio de precios"
    else:
        table_title = "Tabla operativa de precios sugeridos"

    story.append(Paragraph(table_title, PDF_STYLES["SectionTitleGreenCenter"]))
    story.append(Spacer(1, 4))

    if not rows:
        empty_box = Table([[
            Paragraph(
                "Este mercado no tiene productos con precio propio válido en el archivo cargado.",
                PDF_STYLES["BodySoft"]
            )
        ]], colWidths=[265 * mm])

        empty_box.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#cbd5e1")),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 12),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ]))
        story.append(empty_box)
    else:
        story.append(build_prices_table(rows, report_type))

    story.append(PageBreak())
    story.append(build_author_books_contact_block())
    story.append(Spacer(1, 8))

    story.append(Spacer(1, 24))
    story.append(Paragraph(
        "Documento generado automáticamente por PrecioDólar LIVE.",
        PDF_STYLES["SmallMuted"]
    ))

    doc.build(story)
    return pdf_path


def generate_market_fx_zip(report_data: dict, market_key: str, fx_key: str) -> Path:
    ensure_data_dir()
    validate_fx_key(fx_key)

    market = find_market(report_data, market_key)
    if not market:
        raise HTTPException(status_code=404, detail=f"No existe el mercado {market_key}.")

    fx_views = market.get("fx_views", {})
    if fx_key not in fx_views:
        raise HTTPException(status_code=404, detail=f"No existe la referencia {fx_key} para {market_key}.")

    company_name = report_data.get("company_name", "Empresa")

    pdf_operativo = generate_market_fx_pdf(report_data, market_key, "operativo", fx_key)
    pdf_ejecutivo = generate_market_fx_pdf(report_data, market_key, "ejecutivo", fx_key)
    pdf_tienda = generate_market_fx_pdf(report_data, market_key, "tienda", fx_key)
    excel_operativo = generate_market_fx_excel(report_data, market_key, fx_key)
    html_desktop = generate_market_fx_html(report_data, market_key, "desktop", fx_key)
    html_mobile = generate_market_fx_html(report_data, market_key, "mobile", fx_key)

    base_filename = sanitize_filename(f"{company_name}_{market_key}_paquete_completo_{fx_key}")
    zip_path = ZIP_DIR / f"{base_filename}.zip"

    files_to_zip = [
        pdf_operativo,
        pdf_ejecutivo,
        pdf_tienda,
        excel_operativo,
        html_desktop,
        html_mobile,
    ]

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zipf:
        for file_path in files_to_zip:
            zipf.write(file_path, arcname=file_path.name)

    return zip_path



def get_market_fx_rows(report_data: dict, market_key: str, fx_key: str):
    validate_fx_key(fx_key)

    market = find_market(report_data, market_key)
    if not market:
        raise HTTPException(status_code=404, detail=f"No existe el mercado {market_key}.")

    fx_views = market.get("fx_views", {})
    if fx_key not in fx_views:
        raise HTTPException(status_code=404, detail=f"No existe la referencia {fx_key} para {market_key}.")

    return market, fx_views.get(fx_key, {}).get("rows", []) or []


def generate_market_fx_woocommerce_json(report_data: dict, market_key: str, fx_key: str) -> Path:
    ensure_data_dir()
    market, rows = get_market_fx_rows(report_data, market_key, fx_key)
    company_name = report_data.get("company_name", "Empresa")

    products = []
    for row in rows:
        product_name = safe_text(row.get("nombre_producto"), "Producto")
        sku = safe_text(row.get("sku"), "")
        price = format_money(row.get("precio_nuevo_usd", 0))

        products.append({
            "name": product_name,
            "type": "simple",
            "sku": sku if sku != "—" else "",
            "regular_price": price,
            "price": price,
            "meta_data": [
                {"key": "precio_viejo_usd", "value": format_money(row.get("precio_viejo_usd", 0))},
                {"key": "cambio_pct", "value": format_pct(row.get("cambio_pct", 0))},
                {"key": "mercado", "value": market.get("market_label", market_key)},
                {"key": "fx_key", "value": fx_key}
            ]
        })

    base_filename = sanitize_filename(f"{company_name}_{market_key}_woocommerce_{fx_key}")
    json_path = ECOMMERCE_DIR / f"{base_filename}.json"

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(products, f, ensure_ascii=False, indent=2)

    return json_path


def generate_market_fx_shopify_csv(report_data: dict, market_key: str, fx_key: str) -> Path:
    ensure_data_dir()
    market, rows = get_market_fx_rows(report_data, market_key, fx_key)
    company_name = report_data.get("company_name", "Empresa")

    base_filename = sanitize_filename(f"{company_name}_{market_key}_shopify_{fx_key}")
    csv_path = ECOMMERCE_DIR / f"{base_filename}.csv"

    fieldnames = [
        "Handle",
        "Title",
        "Variant SKU",
        "Variant Price",
        "Variant Compare At Price",
        "Status",
        "Tags"
    ]

    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()

        for row in rows:
            product_name = safe_text(row.get("nombre_producto"), "Producto")
            sku = safe_text(row.get("sku"), "")
            handle_source = sku if sku != "—" else product_name

            writer.writerow({
                "Handle": sanitize_filename(handle_source).lower(),
                "Title": product_name,
                "Variant SKU": sku if sku != "—" else "",
                "Variant Price": format_money(row.get("precio_nuevo_usd", 0)),
                "Variant Compare At Price": format_money(row.get("precio_viejo_usd", 0)),
                "Status": "active",
                "Tags": f"PrecioDolarLIVE,{market.get('market_label', market_key)},{fx_key}"
            })

    return csv_path


def generate_market_fx_ecommerce_zip(report_data: dict, market_key: str, fx_key: str) -> Path:
    ensure_data_dir()
    validate_fx_key(fx_key)

    company_name = report_data.get("company_name", "Empresa")
    woocommerce_json = generate_market_fx_woocommerce_json(report_data, market_key, fx_key)
    shopify_csv = generate_market_fx_shopify_csv(report_data, market_key, fx_key)

    base_filename = sanitize_filename(f"{company_name}_{market_key}_ecommerce_{fx_key}")
    zip_path = ZIP_DIR / f"{base_filename}.zip"

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zipf:
        zipf.write(woocommerce_json, arcname=woocommerce_json.name)
        zipf.write(shopify_csv, arcname=shopify_csv.name)

    return zip_path


from fastapi.responses import RedirectResponse

@app.get("/")
def root():
    return RedirectResponse(url="/upload-page")


@app.get("/upload-page")
def upload_page():
    return FileResponse("frontend/input-landing.html")


@app.get("/report-page")
def report_page():
    return FileResponse("frontend/index.html")

from fastapi.responses import FileResponse

@app.get("/pago-bs")
def pago_bs_page():
    return FileResponse("frontend/pago-bs.html")

@app.get("/fx")
def fx_demo():
    snapshot = get_real_fx_snapshot()
    return {
        "as_of_date": snapshot["as_of_date"],
        "summary": get_fx_summary(snapshot)
    }


@app.get("/fx/check")
def fx_check():
    partial = get_partial_real_fx_sources()

    return build_fx_check_response(
        raw_sources=partial["sources"],
        as_of_date=partial["as_of_date"]
    )


@app.post("/fx/complete")
def fx_complete(payload: dict = Body(...)):
    partial = get_partial_real_fx_sources()
    manual_sources = payload.get("sources", {})

    merged_sources = validate_manual_completion(
        auto_sources=partial["sources"],
        manual_sources=manual_sources
    )

    return build_fx_complete_response(
        merged_sources=merged_sources,
        as_of_date=partial["as_of_date"]
    )


@app.post("/upload")
async def upload_excel(
    request: Request,
    response: Response,
    file: UploadFile = File(...),
    fx_snapshot_json: str = Form(...),
    market_source: str = Form("monitor"),
    period_key: str = Form("d7"),
    email: str | None = Form(None)
):

    email = normalize_email(email) or get_email_from_request(request)
    user = get_access_user(email)

    client_id = get_or_create_client_id(request, response)
    access_data = load_access_control()
    access_client = get_or_create_access_client(access_data, client_id)
    save_access_control(access_data)

    if user is not None:
        if not is_user_access_active(user):
            return trial_expired_response()
    else:
        trial_ok = not is_blocked(access_client)
        if not trial_ok:
            return trial_expired_response()

    try:
        content = await file.read()
        payload = read_excel_payload(content)
        old_prices_hash = build_old_prices_hash_from_payload(payload)
    except Exception as exc:
        print("ERROR_UPLOAD_EXCEL:", repr(exc))
        raise HTTPException(
            status_code=400,
            detail=f"No se pudo leer el Excel: {str(exc)}"
        )
    access_status = register_unique_file_usage(client_id, file.filename, old_prices_hash)

    fx_snapshot = json.loads(fx_snapshot_json)
    fx_snapshot = to_pricing_snapshot({
        key: {"value": value, "status": "manual"}
        for key, value in fx_snapshot.items()
    })

    source_mapping = {
        "bcv": "bcv",
        "monitor": "monitor",
        "promedio": "promedio",
        "compuesto": "compuesto",
        "binance": "binance",
        "usdt": "usdt",
        "dolartoday": "dolartoday"
    }

    selected_source = source_mapping.get(market_source, "monitor")

    real_snapshot = get_real_fx_snapshot()

    fx_pair = get_reference_pair(
        snapshot=real_snapshot,
        market_source="oficial" if selected_source == "bcv" else "paralelo",
        period_key=period_key
    )

    fx_pair["tcbc_t"] = real_snapshot.get("bcv", fx_snapshot.get("bcv", 0))
    fx_pair["tcm_t"] = fx_snapshot[selected_source]

    mercados_resumen = []
    report_markets = []

    for idx, mercado in enumerate(payload["mercados"], start=1):
        df = mercado["productos"]

        df_resultado_base, fx_factor_base = apply_pricing_engine(
            df=df.copy(),
            mercado=mercado,
            tcm_t=fx_pair["tcm_t"],
            tcbc_t=fx_pair["tcbc_t"],
            tcm_t_1=fx_pair["tcm_t_1"],
            tcbc_t_1=fx_pair["tcbc_t_1"]
        )

        mercados_resumen.append({
            "sheet_name": mercado["sheet_name"],
            "ciudad": mercado["ciudad"],
            "region": mercado["region"],
            "fx_factor": round(fx_factor_base, 6),
            "total_productos": len(df_resultado_base)
        })

        fx_views = {}

        for fx_key in ["compuesto", "monitor", "binance", "usdt", "dolartoday"]:
            fx_pair_loop = get_reference_pair(
                snapshot=real_snapshot,
                market_source="paralelo",
                period_key=period_key
            )

            if fx_key == "compuesto":
                tcm_today = fx_snapshot.get("compuesto", fx_snapshot.get("promedio", fx_snapshot.get("monitor")))
            else:
                tcm_today = fx_snapshot.get(fx_key, fx_snapshot.get("monitor"))

            tcbc_today = fx_snapshot["bcv"]

            df_resultado, fx_factor = apply_pricing_engine(
                df=df.copy(),
                mercado=mercado,
                tcm_t=tcm_today,
                tcbc_t=tcbc_today,
                tcm_t_1=fx_pair_loop["tcm_t_1"],
                tcbc_t_1=fx_pair_loop["tcbc_t_1"]
            )

            rows = []
            for _, row in df_resultado.iterrows():
                rows.append({
                    "nombre_producto": row.get("nombre_producto", ""),
                    "sku": row.get("sku", ""),
                    "unidad": row.get("unidad_presentacion", ""),
                    "precio_viejo_usd": safe_float(row.get("precio_propio_usd", 0)),
                    "precio_nuevo_usd": safe_float(row.get("precio_sugerido_usd", 0)),
                    "cambio_pct": safe_float(row.get("variacion_precio_pct", 0)),
                    "competidor_lider": safe_float(row.get("precio_lider_usd", 0)),
                    "competidor_intermedio": safe_float(row.get("precio_intermedio_usd", 0)),
                    "competidor_economico": safe_float(row.get("precio_economico_usd", 0)),
                    "peso_competencia": safe_float(row.get("peso_competencia_final", 0)),
                    "peso_riesgo": safe_float(row.get("peso_riesgo_cambiario_final", 0)),
                    "senal": infer_signal(row.get("variacion_precio_pct", 0))
                })

            fx_views[fx_key] = {"rows": rows}

        market_key = f"mercado_{idx}"
        report_markets.append({
            "market_key": market_key,
            "market_label": f"Mercado {idx} · {mercado['ciudad'] or mercado['region']}",
            "city": mercado["ciudad"] or mercado["region"],
            "downloads": build_market_downloads(market_key),
            "fx_views": fx_views
        })

    tcm = fx_pair["tcm_t"]
    tcbc = fx_pair["tcbc_t"]
    brecha = ((tcm / tcbc) - 1) * 100 if tcbc else 0

    company_name = payload.get("empresa") or payload.get("company_name") or "Empresa"

    available = [v for v in fx_snapshot.values() if v not in [None, 0, ""]]
    total = len(fx_snapshot)

    fx_integrity = round(len(available) / total, 2) if total > 0 else 0

    if fx_integrity == 1:
        fx_mode = "full_auto"
    elif fx_snapshot.get("bcv") and not fx_snapshot.get("monitor"):
        fx_mode = "bcv_only"
    else:
       fx_mode = "partial_auto"
    
    report_payload = {
        "company_name": company_name,
        "report_date": real_snapshot["as_of_date"],
        "fx_meta": {
            "fx_integrity": fx_integrity,
            "fx_mode": fx_mode
        },
        "fx_kpis": {
            "compuesto": round(tcm, 2),
            "bcv": round(tcbc, 2),
            "brecha": round(brecha, 1)
        },
        "fx_summary_cards": {
            "compuesto": {
                "valor": round(fx_snapshot.get("compuesto", fx_snapshot.get("promedio", 0)), 2),
                "d7": 5.0,
                "m1": 12.1,
                "m3": 26.9,
                "y1": 79.2
            },
            "binance": {
                "valor": round(fx_snapshot.get("binance", 0), 2),
                "d7": 5.1,
                "m1": 12.4,
                "m3": 27.5,
                "y1": 81.0
            }
        },
        "fx_boxes": {
            "bcv": {
                "valor": round(fx_snapshot.get("bcv", 0), 2),
                "var": 1.9
            },
            "monitor": {
                "valor": round(fx_snapshot.get("monitor", 0), 2),
                "var": 5.2
            },
            "compuesto": {
                "valor": round(fx_snapshot.get("compuesto", fx_snapshot.get("promedio", 0)), 2),
                "var": 5.0
            },
            "binance": {
                "valor": round(fx_snapshot.get("binance", 0), 2),
                "var": 5.1
            },
            "usdt": {
                "valor": round(fx_snapshot.get("usdt", 0), 2),
                "var": 4.9
            },
            "dolartoday": {
                "valor": round(fx_snapshot.get("dolartoday", 0), 2),
                "var": 5.8
            }
        },
        "markets": report_markets
    }

    save_last_report(report_payload, client_id=client_id, email=email)

    return {
        "filename": file.filename,
        "total_mercados_validos": len(payload["mercados"]),
        "mercados": mercados_resumen,
        "persisted": True,
        "json_path": str(get_client_last_report_path(client_id)),
        "access_control": access_status
    }


@app.get("/access")
def access_login_page():
    return render_access_login_page()


@app.post("/access/login")
def access_login(request: Request, email: str = Form(...)):
    normalized_email = normalize_email(email)
    if not normalized_email:
        return render_access_login_page("Email inválido.")

    user = get_access_user(normalized_email)
    if not is_user_access_active(user):
        return render_access_login_page("Ese email no tiene acceso activo. Contacta al administrador para activarlo.")

    response = RedirectResponse(url="/", status_code=303)
    bind_email_to_client(request, response, normalized_email)
    return response


@app.get("/access/status")
def access_status(request: Request):
    email = get_email_from_request(request)
    user = get_access_user(email)
    print("DEBUG EMAIL:", email)
    print("DEBUG USER:", user)
    print("DEBUG USER_OK:", is_user_access_active(user))
    return build_user_access_status(user)


@app.get("/admin")
def admin_page(request: Request):
    return render_admin_page(request)


@app.post("/admin/login")
def admin_login(username: str = Form(...), password: str = Form(...)):
    if username != ADMIN_USERNAME or password != ADMIN_PASSWORD:
        return HTMLResponse("""<!DOCTYPE html><html lang="es"><head><meta charset="utf-8"><meta http-equiv="refresh" content="1;url=/admin"></head><body>Credenciales inválidas.</body></html>""", status_code=401)

    response = RedirectResponse(url="/admin", status_code=303)
    response.set_cookie(
        key=ADMIN_SESSION_COOKIE_NAME,
        value=ADMIN_SESSION_VALUE,
        max_age=60 * 60 * 12,
        httponly=True,
        samesite="lax",
        secure=IS_PROD
    )
    return response


@app.post("/admin/logout")
def admin_logout():
    response = RedirectResponse(url="/admin", status_code=303)
    response.delete_cookie(ADMIN_SESSION_COOKIE_NAME)
    return response


@app.get("/admin/users")
def admin_users(request: Request):
    require_admin_session(request)
    data = load_access_control()
    return {"users": [normalize_access_user(email, user) for email, user in data.get("users", {}).items()]}


@app.post("/admin/users/upsert")
def admin_users_upsert(
    request: Request,
    email: str = Form(...),
    full_name: str | None = Form(None),
    plan: str = Form("basic"),
    duration_days: int = Form(30),
    active: str = Form("true")
):
    require_admin_session(request)
    user = upsert_access_user(
        email=email,
        plan=plan,
        access_until=calculate_access_until(duration_days=duration_days),
        active=str(active).lower() == "true",
        source="manual",
        full_name=full_name
    )
    return render_admin_page(request, f"Usuario guardado: {user['email']}")


@app.post("/admin/users/json")
def admin_users_json(request: Request, payload: dict = Body(...)):
    require_admin_session(request)
    access_until = payload.get("access_until") or calculate_access_until(
        duration_days=payload.get("duration_days"),
        duration=payload.get("duration")
    )
    return upsert_access_user(
        email=payload.get("email"),
        plan=payload.get("plan", "basic"),
        access_until=access_until,
        active=payload.get("active", True),
        source=payload.get("source", "manual")
    )


@app.post("/admin/users/active")
def admin_users_active(request: Request, payload: dict = Body(...)):
    require_admin_session(request)
    return set_user_active(payload.get("email"), bool(payload.get("active", True)))


def require_stripe_config() -> None:
    missing = []
    if stripe is None:
        missing.append("stripe package")
    if not STRIPE_SECRET_KEY:
        missing.append("STRIPE_SECRET_KEY")
    if not STRIPE_WEBHOOK_SECRET:
        missing.append("STRIPE_WEBHOOK_SECRET")
    if not STRIPE_PRICE_BASIC:
        missing.append("STRIPE_PRICE_BASIC")
    if not STRIPE_PRICE_PREMIUM:
        missing.append("STRIPE_PRICE_PREMIUM")
    if not PUBLIC_BASE_URL:
        missing.append("PUBLIC_BASE_URL")

    if missing:
        raise HTTPException(
            status_code=503,
            detail=f"Stripe no está configurado: faltan {', '.join(missing)}."
        )

    stripe.api_key = STRIPE_SECRET_KEY


def get_stripe_price_for_plan(plan: str) -> str:
    if plan == "basic":
        return STRIPE_PRICE_BASIC
    if plan == "premium":
        return STRIPE_PRICE_PREMIUM
    raise HTTPException(status_code=400, detail="Plan inválido. Usa 'basic' o 'premium'.")


def get_plan_from_stripe_price(price_id: str | None) -> str | None:
    if price_id == STRIPE_PRICE_BASIC:
        return "basic"
    if price_id == STRIPE_PRICE_PREMIUM:
        return "premium"
    return None


def stripe_access_until() -> str:
    return (datetime.now(timezone.utc) + timedelta(days=32)).isoformat()


def get_user_by_stripe_customer_or_subscription(customer_id: str | None = None, subscription_id: str | None = None):
    data = load_access_control()
    users = data.setdefault("users", {})

    for email, user in users.items():
        if not isinstance(user, dict):
            continue
        if customer_id and user.get("stripe_customer_id") == customer_id:
            return data, users, email, normalize_access_user(email, user)
        if subscription_id and user.get("stripe_subscription_id") == subscription_id:
            return data, users, email, normalize_access_user(email, user)

    return data, users, None, None


def save_stripe_fields(email: str, customer_id: str | None = None, subscription_id: str | None = None) -> dict:
    normalized_email = normalize_email(email)
    if not normalized_email:
        raise HTTPException(status_code=400, detail="Email Stripe inválido.")

    data = load_access_control()
    users = data.setdefault("users", {})
    user = normalize_access_user(normalized_email, users.get(normalized_email))

    if customer_id:
        user["stripe_customer_id"] = customer_id
    if subscription_id:
        user["stripe_subscription_id"] = subscription_id

    user["updated_at"] = utc_now_iso()
    users[normalized_email] = user
    save_access_control(data)
    return user


def set_stripe_user_active_by_customer_or_subscription(customer_id: str | None = None, subscription_id: str | None = None, active: bool = False) -> dict | None:
    data, users, email, user = get_user_by_stripe_customer_or_subscription(customer_id, subscription_id)
    if not email or not user:
        return None

    user["active"] = bool(active)
    user["updated_at"] = utc_now_iso()
    users[email] = user
    save_access_control(data)
    return user


def get_subscription_price_id(subscription_obj) -> str | None:
    try:
        items = subscription_obj.get("items", {}).get("data", [])
        if items:
            return items[0].get("price", {}).get("id")
    except Exception:
        return None
    return None


def get_invoice_price_id(invoice_obj) -> str | None:
    try:
        lines = invoice_obj.get("lines", {}).get("data", [])
        if lines:
            return lines[0].get("price", {}).get("id")
    except Exception:
        return None
    return None


def get_email_from_stripe_customer(customer_id: str | None) -> str | None:
    if not customer_id:
        return None

    try:
        customer = stripe.Customer.retrieve(customer_id)
        return normalize_email(customer.get("email"))
    except Exception:
        return None


def mark_stripe_event(data: dict, event_id: str, processed: bool, note: str = "") -> None:
    fresh_data = load_access_control()
    events = fresh_data.setdefault("stripe_events", [])

    for item in events:
        if item.get("id") == event_id:
            item["processed"] = bool(processed)
            item["processed_at"] = utc_now_iso()
            if note:
                item["note"] = note
            save_access_control(fresh_data)
            return

    events.append({
        "id": event_id,
        "type": None,
        "received_at": utc_now_iso(),
        "processed": bool(processed),
        "processed_at": utc_now_iso(),
        "note": note
    })
    save_access_control(fresh_data)


def stripe_event_already_processed(event_id: str) -> bool:
    data = load_access_control()
    for item in data.get("stripe_events", []):
        if item.get("id") == event_id and bool(item.get("processed", False)):
            return True
    return False


def register_stripe_event_received(event: dict) -> dict:
    data = load_access_control()
    events = data.setdefault("stripe_events", [])

    event_id = event.get("id")
    event_type = event.get("type")

    for item in events:
        if item.get("id") == event_id:
            return data

    events.append({
        "id": event_id,
        "type": event_type,
        "received_at": utc_now_iso(),
        "processed": False
    })
    save_access_control(data)
    return data


def process_checkout_session_completed(session: dict) -> bool:
    if session.get("mode") not in ["payment", "subscription"]:
        return False

    session_id = session.get("id")
    if not session_id:
        return False

    customer_details = session.get("customer_details") or {}
    full_name = customer_details.get("name")

    email = normalize_email(customer_details.get("email"))
    if not email:
        email = normalize_email(session.get("customer_email"))

    customer_id = session.get("customer")
    if not email and customer_id:
        try:
            customer = stripe.Customer.retrieve(customer_id)
            email = normalize_email(customer.get("email"))
        except Exception:
            email = None

    price_id = None
    amount_total = session.get("amount_total")

    try:
        line_items = stripe.checkout.Session.list_line_items(session_id, limit=10)
        items = getattr(line_items, "data", []) or []
        if items:
            item = items[0]
            price = getattr(item, "price", None)

            if price is None and isinstance(item, dict):
                price = item.get("price")

            if price is not None:
                price_id = getattr(price, "id", None)
                if price_id is None and isinstance(price, dict):
                    price_id = price.get("id")
    except Exception as exc:
        price_id = None

    plan = None
    if price_id == STRIPE_PRICE_BASIC:
        plan = "basic"
    elif price_id == STRIPE_PRICE_PREMIUM:
        plan = "premium"

    if not plan:
        if amount_total == 4900:
            plan = "basic"
        elif amount_total == 9900:
            plan = "premium"

    
    if not email or not plan:
        return False

    user = upsert_access_user(
    email=email,
    plan=plan,
    access_until=calculate_access_until(duration_days=30),
    active=True,
    source="stripe",
    full_name=full_name
)


    saved_user = get_access_user(email)

    return bool(saved_user and saved_user.get("active") is True)


def process_invoice_paid(invoice: dict) -> bool:
    customer_id = invoice.get("customer")
    subscription_id = invoice.get("subscription")
    price_id = get_invoice_price_id(invoice)
    plan = get_plan_from_stripe_price(price_id)

    if not plan and subscription_id:
        try:
            subscription = stripe.Subscription.retrieve(subscription_id)
            price_id = get_subscription_price_id(subscription)
            plan = get_plan_from_stripe_price(price_id)
        except Exception:
            plan = None

    email = normalize_email(invoice.get("customer_email")) or get_email_from_stripe_customer(customer_id)

    if not email or plan not in VALID_ACCESS_PLANS:
        return False

    upsert_access_user(
        email=email,
        plan=plan,
        access_until=stripe_access_until(),
        active=True,
        source="stripe"
    )
    save_stripe_fields(
        email=email,
        customer_id=customer_id,
        subscription_id=subscription_id
    )

    return True


def process_subscription_updated(subscription: dict) -> bool:
    customer_id = subscription.get("customer")
    subscription_id = subscription.get("id")
    status = subscription.get("status")
    price_id = get_subscription_price_id(subscription)
    plan = get_plan_from_stripe_price(price_id)

    email = None
    data, users, existing_email, existing_user = get_user_by_stripe_customer_or_subscription(customer_id, subscription_id)
    if existing_email:
        email = existing_email
    else:
        email = get_email_from_stripe_customer(customer_id)

    if not email:
        return False

    active_statuses = {"active", "trialing"}
    inactive_statuses = {"past_due", "unpaid", "canceled", "incomplete_expired"}

    if status in active_statuses and plan in VALID_ACCESS_PLANS:
        user = upsert_access_user(
            email=email,
            plan=plan,
            access_until=stripe_access_until(),
            active=True,
            source="stripe"
        )

        save_stripe_fields(email=email, customer_id=customer_id, subscription_id=subscription_id)

    
        saved_user = get_access_user(email)
    
        return bool(saved_user and saved_user.get("active") is True)

    if status in inactive_statuses:
        set_stripe_user_active_by_customer_or_subscription(customer_id, subscription_id, active=False)
        return True

    return False


def process_subscription_deleted(subscription: dict) -> bool:
    user = set_stripe_user_active_by_customer_or_subscription(
        customer_id=subscription.get("customer"),
        subscription_id=subscription.get("id"),
        active=False
    )
    return user is not None


@app.post("/stripe/create-checkout-session")
def stripe_create_checkout_session(payload: dict = Body(...)):
    require_stripe_config()

    email = normalize_email(payload.get("email"))
    plan = payload.get("plan")

    if not email:
        raise HTTPException(status_code=400, detail="Email inválido u obligatorio.")
    if plan not in VALID_ACCESS_PLANS:
        raise HTTPException(status_code=400, detail="Plan inválido. Usa 'basic' o 'premium'.")

    price_id = get_stripe_price_for_plan(plan)

    try:
        session = stripe.checkout.Session.create(
            mode="subscription",
            payment_method_types=["card"],
            line_items=[{"price": price_id, "quantity": 1}],
            customer_email=email,
            metadata={
                "email": email,
                "plan": plan
            },
            success_url=f"{PUBLIC_BASE_URL}/access?stripe=success",
            cancel_url=f"{PUBLIC_BASE_URL}/access?stripe=cancel"
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo crear la sesión de Stripe: {str(exc)}")

    return {"url": session.url}


@app.post("/stripe/create-portal-session")
def stripe_create_portal_session(payload: dict = Body(...)):
    require_stripe_config()

    email = normalize_email(payload.get("email"))
    if not email:
        raise HTTPException(status_code=400, detail="Email inválido u obligatorio.")

    user = get_access_user(email)
    if not isinstance(user, dict) or not user.get("stripe_customer_id"):
        raise HTTPException(status_code=404, detail="No existe customer de Stripe para este usuario.")

    try:
        portal = stripe.billing_portal.Session.create(
            customer=user["stripe_customer_id"],
            return_url=f"{PUBLIC_BASE_URL}/access"
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo crear el portal de Stripe: {str(exc)}")

    return {"url": portal.url}


def process_invoice_payment_failed(invoice: dict) -> bool:
    customer_id = invoice.get("customer")
    subscription_id = invoice.get("subscription")

    user = set_stripe_user_active_by_customer_or_subscription(
        customer_id=customer_id,
        subscription_id=subscription_id,
        active=False
    )

    return user is not None


@app.post("/stripe/webhook")
async def stripe_webhook(request: Request):
    require_stripe_config()

    payload = await request.body()
    sig_header = request.headers.get("stripe-signature")

    try:
        event = stripe.Webhook.construct_event(
            payload,
            sig_header,
            STRIPE_WEBHOOK_SECRET
        )
    except Exception:
        raise HTTPException(status_code=400, detail="Firma Stripe inválida.")

    event = event._to_dict_recursive()

    event_id = event.get("id")
    event_type = event.get("type")


    if event_id and stripe_event_already_processed(event_id):
        return {"received": True, "duplicate": True}

    data = register_stripe_event_received(event)

    processed = False

    try:
        obj = event.get("data", {}).get("object", {})

        if event_type == "checkout.session.completed":
            processed = process_checkout_session_completed(obj)
        elif event_type == "invoice.paid":
            processed = process_invoice_paid(obj)
        elif event_type == "invoice.payment_failed": 
            processed = process_invoice_payment_failed(obj)
        elif event_type == "customer.subscription.updated":
            processed = process_subscription_updated(obj)
        elif event_type == "customer.subscription.deleted":
            processed = process_subscription_deleted(obj)
        else:
            processed = False

        if event_id:
            mark_stripe_event(data, event_id, processed, "Procesado" if processed else "Evento recibido sin acción aplicable.")

    except Exception as exc:
        if event_id:
            mark_stripe_event(data, event_id, False, f"Error procesando evento: {str(exc)}")
        raise HTTPException(status_code=500, detail=f"Error procesando webhook Stripe: {str(exc)}")

    return {"received": True, "processed": processed}


@app.post("/dev/mark-paid-basic")
def dev_mark_paid_basic(request: Request, response: Response):
    require_admin_token(request)
    client_id = get_or_create_client_id(request, response)
    return mark_client_paid(client_id, "basic", True)


@app.post("/dev/mark-paid-ecommerce")
def dev_mark_paid_ecommerce(request: Request, response: Response):
    require_admin_token(request)
    client_id = get_or_create_client_id(request, response)
    return mark_client_paid(client_id, "ecommerce", True)


@app.get("/admin/access-control")
def admin_access_control(request: Request):
    require_admin_token(request)
    return load_access_control()


@app.post("/admin/grant-access")
def admin_grant_access(request: Request, payload: dict = Body(...)):
    require_admin_token(request)
    client_id = payload.get("client_id")
    plan = payload.get("plan")

    if not client_id:
        raise HTTPException(status_code=400, detail="client_id es obligatorio.")

    return mark_client_paid(client_id, plan, True)


@app.post("/admin/mark-paid")
def admin_mark_paid(request: Request, payload: dict = Body(...)):
    require_admin_token(request)
    client_id = payload.get("client_id")
    plan = payload.get("plan")

    if not client_id:
        raise HTTPException(status_code=400, detail="client_id es obligatorio.")

    return mark_client_paid(client_id, plan, True)


@app.post("/admin/revoke-access")
def admin_revoke_access(request: Request, payload: dict = Body(...)):
    require_admin_token(request)
    client_id = payload.get("client_id")
    plan = payload.get("plan")

    if not client_id:
        raise HTTPException(status_code=400, detail="client_id es obligatorio.")

    return mark_client_paid(client_id, plan, False)


@app.get("/report/data")
def report_data(request: Request, response: Response):
    access_response = require_basic_access(request)
    if access_response:
        return access_response

    client_id = get_or_create_client_id(request, response)
    email = get_email_from_request(request)

    if email:
        report = load_last_report(None, email=email)
    else:
        report = load_last_report(client_id)

    if report is None:
        return JSONResponse(
            status_code=404,
            content={
                "error": "No hay reporte persistido todavía para este cliente o email.",
                "email": email,
                "client_id": client_id
            }
        )

    return report

@app.post("/fx/save")
def save_fx(data: dict = Body(...)):
    today = datetime.now().strftime("%Y-%m-%d")
    history = load_fx_history()

    entry = normalize_fx_entry(data, today)
    history, status = upsert_fx_history_entry(history, entry)

    save_fx_history(history)

    return {"status": status, "entry": entry}

@app.get("/fx/history")
def get_fx_history(period: str = "7d"):

    history = get_fx_history_cached()

    days_map = {
        "7d": 7,
        "1m": 30,
        "3m": 90,
        "1a": 365
    }

    days = days_map.get(period, 7)
    cutoff = datetime.now() - timedelta(days=days)

    filtered = []

    for h in history:
        try:
            raw_date = str(h.get("date", "")).strip()[:10]
            dt = datetime.strptime(raw_date, "%Y-%m-%d")

            if dt >= cutoff:
                h["date"] = raw_date
                filtered.append(h)

        except Exception:
            continue

    return filtered
    
@app.get("/fx/pricing-context-test")
def fx_pricing_context_test(reference: str = "compuesto"):
    try:
        return get_pricing_fx_context(
            selected_reference=reference,
            force_refresh=True
        )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "error": str(e),
                "type": type(e).__name__
            }
        )

@app.get("/report/pdf/{market_key}/{report_type}/{fx_key}")
def report_pdf(request: Request, market_key: str, report_type: str, fx_key: str):
    access_response = require_basic_access(request)
    if access_response:
        return access_response

    report_data = get_report_data_or_raise(request)
    pdf_path = generate_market_fx_pdf(report_data, market_key, report_type, fx_key)

    return FileResponse(
        path=str(pdf_path),
        media_type="application/pdf",
        filename=pdf_path.name
    )

@app.get("/report/excel/{market_key}/{fx_key}")
def report_excel(request: Request, market_key: str, fx_key: str):
    access_response = require_basic_access(request)
    if access_response:
        return access_response

    report_data = get_report_data_or_raise(request)
    excel_path = generate_market_fx_excel(report_data, market_key, fx_key)

    return FileResponse(
        path=str(excel_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=excel_path.name
    )


@app.get("/report/html/{market_key}/{view_type}/{fx_key}")
def report_html(request: Request, market_key: str, view_type: str, fx_key: str):
    access_response = require_basic_access(request)
    if access_response:
        return access_response

    report_data = get_report_data_or_raise(request)
    html_path = generate_market_fx_html(report_data, market_key, view_type, fx_key)

    return FileResponse(
        path=str(html_path),
        media_type="text/html",
        filename=html_path.name
    )


@app.get("/report/zip/{market_key}/{fx_key}")
def report_zip(request: Request, market_key: str, fx_key: str):
    access_response = require_basic_access(request)
    if access_response:
        return access_response

    report_data = get_report_data_or_raise(request)
    zip_path = generate_market_fx_zip(report_data, market_key, fx_key)

    return FileResponse(
        path=str(zip_path),
        media_type="application/zip",
        filename=zip_path.name
    )


@app.get("/report/woocommerce/{market_key}/{fx_key}")
def report_woocommerce(request: Request, market_key: str, fx_key: str):
    access_response = require_ecommerce_access(request)
    if access_response:
        return access_response

    report_data = get_report_data_or_raise(request)
    json_path = generate_market_fx_woocommerce_json(report_data, market_key, fx_key)

    return FileResponse(
        path=str(json_path),
        media_type="application/json",
        filename=json_path.name
    )


@app.get("/report/shopify/{market_key}/{fx_key}")
def report_shopify(request: Request, market_key: str, fx_key: str):
    access_response = require_ecommerce_access(request)
    if access_response:
        return access_response

    report_data = get_report_data_or_raise(request)
    csv_path = generate_market_fx_shopify_csv(report_data, market_key, fx_key)

    return FileResponse(
        path=str(csv_path),
        media_type="text/csv",
        filename=csv_path.name
    )


@app.get("/report/ecommerce-zip/{market_key}/{fx_key}")
def report_ecommerce_zip(request: Request, market_key: str, fx_key: str):
    access_response = require_ecommerce_access(request)
    if access_response:
        return access_response

    report_data = get_report_data_or_raise(request)
    zip_path = generate_market_fx_ecommerce_zip(report_data, market_key, fx_key)

    return FileResponse(
        path=str(zip_path),
        media_type="application/zip",
        filename=zip_path.name
    )

@app.post("/fx/update-daily")
async def update_daily_fx(request: Request):
    try:
        try:
            body = await request.json()
        except Exception:
            raw = await request.body()
            return {
                "ok": False,
                "error": f"Invalid JSON received → {raw.decode('utf-8')[:200]}"
            }

        if body.get("secret") != CRONSECRET:
            raise HTTPException(status_code=403, detail="Forbidden")

        today = datetime.now().strftime("%Y-%m-%d")

        history = load_fx_history(force_refresh=True)

        if fx_entry_exists(history, today):
            return {
                "ok": True,
                "status": "duplicate",
                "date": today
            }

        data = build_flat_fx_values()

        if not data:
            raise Exception("No FX data returned from extractors")

        entry = {
            "date": today,
            "bcv": round(safe_float(data.get("bcv")), 2),
            "monitor": round(safe_float(data.get("monitor")), 2),
            "binance": round(safe_float(data.get("binance")), 2),
            "usdt": round(safe_float(data.get("usdt")), 2),
            "dolartoday": round(safe_float(data.get("dolartoday")), 2),
        }

        entry["compuesto"] = round(calculate_compuesto(entry), 2)

        status = save_fx_entry(entry)

        return {
            "ok": True,
            "status": status,
            "date": today
        }

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": str(e),
                "type": type(e).__name__
            }
        )
